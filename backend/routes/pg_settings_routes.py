"""
PostgreSQL System Settings & Reports Routes
Migrated from MongoDB to PostgreSQL
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func, desc, and_
import uuid
import json
import io

from database import (
    get_postgres_session, SystemSetting, AuditLog, User,
    PurchaseOrder, PurchaseOrderItem, Project, BudgetCategory, Supplier, MaterialRequest
)

# Create router
pg_settings_router = APIRouter(prefix="/api/pg", tags=["PostgreSQL Settings & Reports"])

# Import auth dependency
from routes.pg_auth_routes import get_current_user_pg, UserRole


# ==================== PYDANTIC MODELS ====================

class SystemSettingUpdate(BaseModel):
    value: str


# ==================== SYSTEM SETTINGS ROUTES ====================

@pg_settings_router.get("/settings")
async def get_system_settings(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get all system settings - procurement manager and general manager"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.GENERAL_MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    result = await session.execute(select(SystemSetting))
    settings = result.scalars().all()
    
    return [
        {
            "id": s.id,
            "key": s.key,
            "value": s.value,
            "description": s.description,
            "updated_by_name": s.updated_by_name,
            "updated_at": s.updated_at.isoformat() if s.updated_at else None
        }
        for s in settings
    ]


@pg_settings_router.get("/settings/{key}")
async def get_system_setting(
    key: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get a single system setting"""
    result = await session.execute(
        select(SystemSetting).where(SystemSetting.key == key)
    )
    setting = result.scalar_one_or_none()
    
    if not setting:
        raise HTTPException(status_code=404, detail="الإعداد غير موجود")
    
    return {
        "key": setting.key,
        "value": setting.value,
        "description": setting.description
    }


@pg_settings_router.put("/settings/{key}")
async def update_system_setting(
    key: str,
    update_data: SystemSettingUpdate,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Update a system setting - procurement manager and general manager"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.GENERAL_MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    result = await session.execute(
        select(SystemSetting).where(SystemSetting.key == key)
    )
    setting = result.scalar_one_or_none()
    
    if not setting:
        raise HTTPException(status_code=404, detail="الإعداد غير موجود")
    
    old_value = setting.value
    setting.value = update_data.value
    setting.updated_by = current_user.id
    setting.updated_by_name = current_user.name
    setting.updated_at = datetime.utcnow()
    
    # Log audit
    audit_log = AuditLog(
        id=str(uuid.uuid4()),
        entity_type="setting",
        entity_id=setting.id,
        action="update",
        changes=json.dumps({"old": old_value, "new": update_data.value}),
        user_id=current_user.id,
        user_name=current_user.name,
        user_role=current_user.role,
        description=f"تحديث الإعداد: {key}"
    )
    session.add(audit_log)
    
    await session.commit()
    
    return {"message": "تم تحديث الإعداد بنجاح"}


@pg_settings_router.post("/settings/init")
async def init_system_settings(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Initialize default system settings - procurement manager only"""
    if current_user.role != UserRole.PROCUREMENT_MANAGER:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    default_settings = [
        {"key": "approval_limit", "value": "20000", "description": "حد الموافقة - المبالغ الأعلى تحتاج موافقة المدير العام"},
        {"key": "company_name", "value": "شركة المشتريات", "description": "اسم الشركة"},
        {"key": "company_address", "value": "", "description": "عنوان الشركة"},
        {"key": "company_phone", "value": "", "description": "هاتف الشركة"},
        {"key": "company_email", "value": "", "description": "البريد الإلكتروني للشركة"},
        {"key": "currency", "value": "ريال سعودي", "description": "العملة المستخدمة"},
        {"key": "vat_rate", "value": "15", "description": "نسبة ضريبة القيمة المضافة"},
    ]
    
    now = datetime.utcnow()
    added = 0
    
    for setting_data in default_settings:
        # Check if exists
        result = await session.execute(
            select(SystemSetting).where(SystemSetting.key == setting_data["key"])
        )
        if result.scalar_one_or_none():
            continue
        
        setting = SystemSetting(
            id=str(uuid.uuid4()),
            key=setting_data["key"],
            value=setting_data["value"],
            description=setting_data["description"],
            updated_by=current_user.id,
            updated_by_name=current_user.name,
            created_at=now
        )
        session.add(setting)
        added += 1
    
    await session.commit()
    
    return {"message": f"تم إضافة {added} إعداد", "added_count": added}


# ==================== REPORTS ROUTES ====================

@pg_settings_router.get("/reports/cost-savings")
async def get_cost_savings_report(
    project_id: Optional[str] = None,
    category_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get cost savings report - available for procurement manager and general manager"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.GENERAL_MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    query = select(PurchaseOrder).where(PurchaseOrder.status.in_(["approved", "printed", "shipped", "delivered"]))
    
    if project_id:
        query = query.where(PurchaseOrder.project_id == project_id)
    if category_id:
        query = query.where(PurchaseOrder.category_id == category_id)
    if start_date:
        query = query.where(PurchaseOrder.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.where(PurchaseOrder.created_at <= datetime.fromisoformat(end_date))
    
    result = await session.execute(query)
    orders = result.scalars().all()
    
    # Group by project
    by_project = {}
    for order in orders:
        project_name = order.project_name
        if project_name not in by_project:
            by_project[project_name] = {
                "project_name": project_name,
                "total_amount": 0,
                "order_count": 0
            }
        by_project[project_name]["total_amount"] += order.total_amount
        by_project[project_name]["order_count"] += 1
    
    # Group by category
    by_category = {}
    for order in orders:
        category_name = order.category_name or "غير مصنف"
        if category_name not in by_category:
            by_category[category_name] = {
                "category_name": category_name,
                "total_amount": 0,
                "order_count": 0
            }
        by_category[category_name]["total_amount"] += order.total_amount
        by_category[category_name]["order_count"] += 1
    
    # Group by supplier
    by_supplier = {}
    for order in orders:
        supplier_name = order.supplier_name
        if supplier_name not in by_supplier:
            by_supplier[supplier_name] = {
                "supplier_name": supplier_name,
                "total_amount": 0,
                "order_count": 0
            }
        by_supplier[supplier_name]["total_amount"] += order.total_amount
        by_supplier[supplier_name]["order_count"] += 1
    
    total_amount = sum(o.total_amount for o in orders)
    
    # Calculate summary data for frontend compatibility
    summary = {
        "total_estimated": total_amount,  # Using actual as estimated for now
        "total_actual": total_amount,
        "total_saving": 0,
        "saving_percent": 0
    }
    
    # Add summary data to by_project for compatibility
    by_project_with_summary = []
    for p in by_project.values():
        by_project_with_summary.append({
            **p,
            "estimated": p["total_amount"],
            "actual": p["total_amount"],
            "saving": 0,
            "saving_percent": 0
        })
    
    # Add summary data to by_category for compatibility
    by_category_with_summary = []
    for c in by_category.values():
        by_category_with_summary.append({
            **c,
            "estimated": c["total_amount"],
            "actual": c["total_amount"],
            "saving": 0,
            "saving_percent": 0
        })
    
    return {
        "total_orders": len(orders),
        "total_amount": total_amount,
        "summary": summary,
        "by_project": by_project_with_summary,
        "by_category": by_category_with_summary,
        "by_supplier": list(by_supplier.values())
    }


@pg_settings_router.get("/reports/dashboard")
async def get_dashboard_stats(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get dashboard statistics"""
    # Total projects
    projects_result = await session.execute(
        select(func.count()).select_from(Project).where(Project.status == "active")
    )
    total_projects = projects_result.scalar() or 0
    
    # Total suppliers
    suppliers_result = await session.execute(
        select(func.count()).select_from(Supplier)
    )
    total_suppliers = suppliers_result.scalar() or 0
    
    # Orders stats
    orders_result = await session.execute(
        select(
            func.count().label('total'),
            func.coalesce(func.sum(PurchaseOrder.total_amount), 0).label('total_amount')
        ).select_from(PurchaseOrder)
    )
    orders_stats = orders_result.first()
    
    # Pending orders
    pending_result = await session.execute(
        select(func.count()).select_from(PurchaseOrder)
        .where(PurchaseOrder.status.in_(["pending_approval", "pending_gm_approval"]))
    )
    pending_orders = pending_result.scalar() or 0
    
    # Delivered orders
    delivered_result = await session.execute(
        select(func.count()).select_from(PurchaseOrder)
        .where(PurchaseOrder.status == "delivered")
    )
    delivered_orders = delivered_result.scalar() or 0
    
    return {
        "total_projects": total_projects,
        "total_suppliers": total_suppliers,
        "total_orders": orders_stats.total if orders_stats else 0,
        "total_amount": float(orders_stats.total_amount) if orders_stats else 0,
        "pending_orders": pending_orders,
        "delivered_orders": delivered_orders
    }


@pg_settings_router.get("/reports/budget")
async def get_budget_report(
    project_id: Optional[str] = None,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get budget report - spending by category and project"""
    
    # Get all budget categories
    categories_result = await session.execute(
        select(BudgetCategory).order_by(BudgetCategory.name)
    )
    categories = categories_result.scalars().all()
    
    # Get all projects
    projects_query = select(Project).where(Project.status == "active")
    if project_id:
        projects_query = projects_query.where(Project.id == project_id)
    projects_result = await session.execute(projects_query.order_by(Project.name))
    projects = projects_result.scalars().all()
    
    budget_data = []
    
    for project in projects:
        project_budget = {
            "project_id": project.id,
            "project_name": project.name,
            "categories": [],
            "total_spent": 0,
            "total_budget": 0
        }
        
        # Get categories for this project
        project_cats_result = await session.execute(
            select(BudgetCategory).where(BudgetCategory.project_id == project.id)
        )
        project_cats = project_cats_result.scalars().all()
        
        for category in project_cats:
            # Sum of approved orders in this category for this project
            spending_result = await session.execute(
                select(func.coalesce(func.sum(PurchaseOrder.total_amount), 0))
                .select_from(PurchaseOrder)
                .where(
                    PurchaseOrder.project_id == project.id,
                    PurchaseOrder.category_id == category.id,
                    PurchaseOrder.status.in_(["approved", "printed", "shipped", "delivered"])
                )
            )
            spent = float(spending_result.scalar() or 0)
            cat_budget = float(category.estimated_budget or 0)
            
            project_budget["categories"].append({
                "category_id": category.id,
                "category_name": category.name,
                "budget": cat_budget,
                "spent": spent,
                "remaining": cat_budget - spent,
                "percentage": round((spent / cat_budget * 100), 1) if cat_budget > 0 else 0
            })
            
            project_budget["total_spent"] += spent
            project_budget["total_budget"] += cat_budget
        
        project_budget["remaining"] = project_budget["total_budget"] - project_budget["total_spent"]
        project_budget["percentage"] = round(
            (project_budget["total_spent"] / project_budget["total_budget"] * 100), 1
        ) if project_budget["total_budget"] > 0 else 0
        
        budget_data.append(project_budget)
    
    # Summary
    total_all_budgets = sum(p["total_budget"] for p in budget_data)
    total_all_spent = sum(p["total_spent"] for p in budget_data)
    
    return {
        "projects": budget_data,
        "summary": {
            "total_budget": total_all_budgets,
            "total_spent": total_all_spent,
            "total_remaining": total_all_budgets - total_all_spent,
            "overall_percentage": round((total_all_spent / total_all_budgets * 100), 1) if total_all_budgets > 0 else 0
        }
    }


@pg_settings_router.get("/reports/budget/export")
async def export_budget_report(
    project_id: Optional[str] = None,
    format: str = "excel",
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تصدير تقرير الميزانية إلى Excel"""
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة Excel غير متوفرة")
    
    # Get budget data
    projects_query = select(Project).where(Project.status == "active")
    if project_id:
        projects_query = projects_query.where(Project.id == project_id)
    projects_result = await session.execute(projects_query.order_by(Project.name))
    projects = projects_result.scalars().all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير الميزانية"
    ws.sheet_view.rightToLeft = True
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    danger_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"📊 تقرير الميزانية - {datetime.now().strftime('%Y-%m-%d')}"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    
    row_num = 3
    total_all_budget = 0
    total_all_spent = 0
    
    for project in projects:
        # Project header
        ws.merge_cells(f'A{row_num}:F{row_num}')
        project_cell = ws.cell(row=row_num, column=1, value=f"🏢 {project.name}")
        project_cell.font = Font(bold=True, size=12)
        project_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        row_num += 1
        
        # Headers
        headers = ['التصنيف', 'الميزانية', 'المصروف', 'المتبقي', 'النسبة %', 'الحالة']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        row_num += 1
        
        # Get categories for project
        cats_result = await session.execute(
            select(BudgetCategory).where(BudgetCategory.project_id == project.id)
        )
        categories = cats_result.scalars().all()
        
        project_total_budget = 0
        project_total_spent = 0
        
        for category in categories:
            # Calculate spending
            spending_result = await session.execute(
                select(func.coalesce(func.sum(PurchaseOrder.total_amount), 0))
                .select_from(PurchaseOrder)
                .where(
                    PurchaseOrder.project_id == project.id,
                    PurchaseOrder.category_id == category.id,
                    PurchaseOrder.status.in_(["approved", "printed", "shipped", "delivered"])
                )
            )
            spent = float(spending_result.scalar() or 0)
            budget = float(category.estimated_budget or 0)
            remaining = budget - spent
            percentage = round((spent / budget * 100), 1) if budget > 0 else 0
            
            # Determine status
            if percentage >= 100:
                status = "تجاوز"
                status_fill = danger_fill
            elif percentage >= 80:
                status = "تحذير"
                status_fill = warning_fill
            else:
                status = "طبيعي"
                status_fill = None
            
            ws.cell(row=row_num, column=1, value=category.name).border = thin_border
            ws.cell(row=row_num, column=2, value=budget).border = thin_border
            ws.cell(row=row_num, column=3, value=spent).border = thin_border
            ws.cell(row=row_num, column=4, value=remaining).border = thin_border
            ws.cell(row=row_num, column=5, value=f"{percentage}%").border = thin_border
            status_cell = ws.cell(row=row_num, column=6, value=status)
            status_cell.border = thin_border
            status_cell.alignment = center_align
            if status_fill:
                status_cell.fill = status_fill
                status_cell.font = Font(bold=True, color="FFFFFF")
            
            row_num += 1
            project_total_budget += budget
            project_total_spent += spent
        
        # Project totals
        ws.cell(row=row_num, column=1, value="المجموع").font = Font(bold=True)
        ws.cell(row=row_num, column=1).border = thin_border
        ws.cell(row=row_num, column=2, value=project_total_budget).font = Font(bold=True)
        ws.cell(row=row_num, column=2).border = thin_border
        ws.cell(row=row_num, column=3, value=project_total_spent).font = Font(bold=True)
        ws.cell(row=row_num, column=3).border = thin_border
        ws.cell(row=row_num, column=4, value=project_total_budget - project_total_spent).font = Font(bold=True)
        ws.cell(row=row_num, column=4).border = thin_border
        
        total_all_budget += project_total_budget
        total_all_spent += project_total_spent
        row_num += 2
    
    # Grand total
    ws.merge_cells(f'A{row_num}:F{row_num}')
    grand_total = ws.cell(row=row_num, column=1, value=f"الإجمالي الكلي: الميزانية {total_all_budget:,.0f} | المصروف {total_all_spent:,.0f} | المتبقي {total_all_budget - total_all_spent:,.0f}")
    grand_total.font = Font(bold=True, size=12)
    grand_total.fill = title_fill
    grand_total.font = Font(bold=True, color="FFFFFF", size=12)
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=budget_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


# ==================== ADVANCED REPORTS ====================

@pg_settings_router.get("/reports/advanced/summary")
async def get_advanced_summary_report(
    project_id: Optional[str] = None,
    engineer_id: Optional[str] = None,
    supervisor_id: Optional[str] = None,
    supplier_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """ملخص تنفيذي شامل - للمدير العام ومدير المشتريات"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.GENERAL_MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    # Build filters for orders
    order_filters = []
    request_filters = []
    
    if project_id:
        order_filters.append(PurchaseOrder.project_id == project_id)
        request_filters.append(MaterialRequest.project_id == project_id)
    if supplier_id:
        order_filters.append(PurchaseOrder.supplier_id == supplier_id)
    if engineer_id:
        request_filters.append(MaterialRequest.engineer_id == engineer_id)
    if supervisor_id:
        request_filters.append(MaterialRequest.supervisor_id == supervisor_id)
    if start_date:
        start = datetime.fromisoformat(start_date)
        order_filters.append(PurchaseOrder.created_at >= start)
        request_filters.append(MaterialRequest.created_at >= start)
    if end_date:
        end = datetime.fromisoformat(end_date)
        order_filters.append(PurchaseOrder.created_at <= end)
        request_filters.append(MaterialRequest.created_at <= end)
    
    # Total Purchase Orders
    orders_query = select(PurchaseOrder)
    if order_filters:
        orders_query = orders_query.where(and_(*order_filters))
    orders_result = await session.execute(orders_query)
    all_orders = orders_result.scalars().all()
    
    # Total Material Requests
    requests_query = select(MaterialRequest)
    if request_filters:
        requests_query = requests_query.where(and_(*request_filters))
    requests_result = await session.execute(requests_query)
    all_requests = requests_result.scalars().all()
    
    # Order statistics
    orders_by_status = {}
    total_order_amount = 0
    for order in all_orders:
        status = order.status or "pending"
        orders_by_status[status] = orders_by_status.get(status, 0) + 1
        if order.status in ["approved", "printed", "shipped", "delivered"]:
            total_order_amount += order.total_amount or 0
    
    # Request statistics
    requests_by_status = {}
    for req in all_requests:
        status = req.status or "pending"
        requests_by_status[status] = requests_by_status.get(status, 0) + 1
    
    # Monthly spending trend (last 6 months)
    monthly_spending = []
    from datetime import timedelta
    now = datetime.utcnow()
    for i in range(5, -1, -1):
        month_start = (now.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1)
        
        month_result = await session.execute(
            select(func.coalesce(func.sum(PurchaseOrder.total_amount), 0))
            .where(
                PurchaseOrder.status.in_(["approved", "printed", "shipped", "delivered"]),
                PurchaseOrder.created_at >= month_start,
                PurchaseOrder.created_at < month_end
            )
        )
        month_total = float(month_result.scalar() or 0)
        monthly_spending.append({
            "month": month_start.strftime("%Y-%m"),
            "month_name": month_start.strftime("%B %Y"),
            "amount": month_total
        })
    
    # Top 5 projects by spending
    top_projects_result = await session.execute(
        select(
            PurchaseOrder.project_name,
            func.sum(PurchaseOrder.total_amount).label("total")
        )
        .where(PurchaseOrder.status.in_(["approved", "printed", "shipped", "delivered"]))
        .group_by(PurchaseOrder.project_name)
        .order_by(desc("total"))
        .limit(5)
    )
    top_projects = [{"name": r[0], "amount": float(r[1] or 0)} for r in top_projects_result.all()]
    
    # Top 5 suppliers by amount
    top_suppliers_result = await session.execute(
        select(
            PurchaseOrder.supplier_name,
            func.sum(PurchaseOrder.total_amount).label("total"),
            func.count(PurchaseOrder.id).label("order_count")
        )
        .where(PurchaseOrder.status.in_(["approved", "printed", "shipped", "delivered"]))
        .group_by(PurchaseOrder.supplier_name)
        .order_by(desc("total"))
        .limit(5)
    )
    top_suppliers = [{"name": r[0], "amount": float(r[1] or 0), "orders": r[2]} for r in top_suppliers_result.all()]
    
    # Spending by category
    by_category_result = await session.execute(
        select(
            PurchaseOrder.category_name,
            func.sum(PurchaseOrder.total_amount).label("total")
        )
        .where(PurchaseOrder.status.in_(["approved", "printed", "shipped", "delivered"]))
        .group_by(PurchaseOrder.category_name)
        .order_by(desc("total"))
    )
    spending_by_category = [{"name": r[0] or "غير مصنف", "amount": float(r[1] or 0)} for r in by_category_result.all()]
    
    return {
        "summary": {
            "total_orders": len(all_orders),
            "total_requests": len(all_requests),
            "total_spending": total_order_amount,
            "approved_orders": orders_by_status.get("approved", 0) + orders_by_status.get("printed", 0) + orders_by_status.get("shipped", 0) + orders_by_status.get("delivered", 0),
            "pending_orders": orders_by_status.get("pending", 0) + orders_by_status.get("pending_gm", 0),
            "rejected_orders": orders_by_status.get("rejected", 0) + orders_by_status.get("rejected_gm", 0)
        },
        "orders_by_status": orders_by_status,
        "requests_by_status": requests_by_status,
        "monthly_spending": monthly_spending,
        "top_projects": top_projects,
        "top_suppliers": top_suppliers,
        "spending_by_category": spending_by_category
    }


@pg_settings_router.get("/reports/advanced/approval-analytics")
async def get_approval_analytics(
    project_id: Optional[str] = None,
    engineer_id: Optional[str] = None,
    supervisor_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تحليل سير الاعتمادات - للمدير العام ومدير المشتريات"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.GENERAL_MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    filters = []
    if project_id:
        filters.append(MaterialRequest.project_id == project_id)
    if engineer_id:
        filters.append(MaterialRequest.engineer_id == engineer_id)
    if supervisor_id:
        filters.append(MaterialRequest.supervisor_id == supervisor_id)
    if start_date:
        filters.append(MaterialRequest.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        filters.append(MaterialRequest.created_at <= datetime.fromisoformat(end_date))
    
    # Get all requests
    query = select(MaterialRequest)
    if filters:
        query = query.where(and_(*filters))
    result = await session.execute(query)
    all_requests = result.scalars().all()
    
    # Approval statistics
    total = len(all_requests)
    approved = len([r for r in all_requests if r.status in ["approved", "po_issued"]])
    rejected = len([r for r in all_requests if r.status in ["rejected", "rejected_engineer"]])
    pending = len([r for r in all_requests if r.status in ["pending", "pending_engineer"]])
    
    # By engineer
    by_engineer = {}
    for req in all_requests:
        engineer = req.engineer_name or "غير محدد"
        if engineer not in by_engineer:
            by_engineer[engineer] = {"approved": 0, "rejected": 0, "pending": 0, "total": 0}
        by_engineer[engineer]["total"] += 1
        if req.status in ["approved", "po_issued"]:
            by_engineer[engineer]["approved"] += 1
        elif req.status in ["rejected", "rejected_engineer"]:
            by_engineer[engineer]["rejected"] += 1
        else:
            by_engineer[engineer]["pending"] += 1
    
    # By supervisor
    by_supervisor = {}
    for req in all_requests:
        supervisor = req.supervisor_name or "غير محدد"
        if supervisor not in by_supervisor:
            by_supervisor[supervisor] = {"approved": 0, "rejected": 0, "pending": 0, "total": 0}
        by_supervisor[supervisor]["total"] += 1
        if req.status in ["approved", "po_issued"]:
            by_supervisor[supervisor]["approved"] += 1
        elif req.status in ["rejected", "rejected_engineer"]:
            by_supervisor[supervisor]["rejected"] += 1
        else:
            by_supervisor[supervisor]["pending"] += 1
    
    # By project
    by_project = {}
    for req in all_requests:
        project = req.project_name or "غير محدد"
        if project not in by_project:
            by_project[project] = {"approved": 0, "rejected": 0, "pending": 0, "total": 0}
        by_project[project]["total"] += 1
        if req.status in ["approved", "po_issued"]:
            by_project[project]["approved"] += 1
        elif req.status in ["rejected", "rejected_engineer"]:
            by_project[project]["rejected"] += 1
        else:
            by_project[project]["pending"] += 1
    
    return {
        "summary": {
            "total_requests": total,
            "approved": approved,
            "rejected": rejected,
            "pending": pending,
            "approval_rate": round((approved / total * 100), 1) if total > 0 else 0,
            "rejection_rate": round((rejected / total * 100), 1) if total > 0 else 0
        },
        "by_engineer": [{"name": k, **v} for k, v in by_engineer.items()],
        "by_supervisor": [{"name": k, **v} for k, v in by_supervisor.items()],
        "by_project": [{"name": k, **v} for k, v in by_project.items()]
    }


@pg_settings_router.get("/reports/advanced/supplier-performance")
async def get_supplier_performance_report(
    supplier_id: Optional[str] = None,
    project_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    item_name: Optional[str] = None,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تقرير أداء الموردين - للمدير العام ومدير المشتريات
    يشمل: الأصناف المرتبطة، أسعار الشراء، الالتزام الزمني
    """
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.GENERAL_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    # Get suppliers
    suppliers_query = select(Supplier)
    if supplier_id:
        suppliers_query = suppliers_query.where(Supplier.id == supplier_id)
    suppliers_result = await session.execute(suppliers_query)
    suppliers = suppliers_result.scalars().all()
    
    performance_data = []
    
    for supplier in suppliers:
        # Orders for this supplier
        orders_query = select(PurchaseOrder).where(PurchaseOrder.supplier_id == supplier.id)
        if project_id:
            orders_query = orders_query.where(PurchaseOrder.project_id == project_id)
        
        # Date filters
        if start_date:
            try:
                start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                orders_query = orders_query.where(PurchaseOrder.created_at >= start)
            except:
                pass
        
        if end_date:
            try:
                end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                orders_query = orders_query.where(PurchaseOrder.created_at <= end)
            except:
                pass
        
        orders_result = await session.execute(orders_query.order_by(desc(PurchaseOrder.created_at)))
        orders = orders_result.scalars().all()
        
        if not orders:
            continue
        
        total_orders = len(orders)
        completed_orders = len([o for o in orders if o.status in ["delivered"]])
        approved_orders = len([o for o in orders if o.status in ["approved", "printed", "shipped", "delivered", "partially_delivered"]])
        total_amount = sum(o.total_amount or 0 for o in orders if o.status in ["approved", "printed", "shipped", "delivered", "partially_delivered"])
        
        # Calculate on-time delivery rate
        on_time_deliveries = 0
        late_deliveries = 0
        pending_late = 0  # Orders that are late but not yet delivered
        
        for order in orders:
            if order.expected_delivery_date:
                try:
                    expected = datetime.strptime(order.expected_delivery_date, "%Y-%m-%d")
                    
                    if order.status == "delivered" and order.delivered_at:
                        if order.delivered_at.replace(tzinfo=None) <= expected:
                            on_time_deliveries += 1
                        else:
                            late_deliveries += 1
                    elif order.status not in ["delivered"] and datetime.now() > expected:
                        pending_late += 1  # Should have been delivered but wasn't
                except:
                    pass
        
        # Get items for this supplier's orders
        items_data = {}
        for order in orders:
            items_result = await session.execute(
                select(PurchaseOrderItem).where(PurchaseOrderItem.order_id == order.id)
            )
            items = items_result.scalars().all()
            
            for item in items:
                # Filter by item name if specified
                if item_name and item_name.lower() not in item.name.lower():
                    continue
                
                if item.name not in items_data:
                    items_data[item.name] = {
                        "name": item.name,
                        "unit": item.unit,
                        "total_quantity": 0,
                        "total_price": 0,
                        "prices": [],
                        "order_count": 0,
                        "min_price": None,
                        "max_price": None
                    }
                
                items_data[item.name]["total_quantity"] += item.quantity
                items_data[item.name]["total_price"] += item.total_price or 0
                items_data[item.name]["order_count"] += 1
                items_data[item.name]["prices"].append({
                    "unit_price": item.unit_price,
                    "quantity": item.quantity,
                    "order_number": order.order_number,
                    "date": order.created_at.strftime("%Y-%m-%d") if order.created_at else None
                })
                
                # Track min/max prices
                if items_data[item.name]["min_price"] is None or item.unit_price < items_data[item.name]["min_price"]:
                    items_data[item.name]["min_price"] = item.unit_price
                if items_data[item.name]["max_price"] is None or item.unit_price > items_data[item.name]["max_price"]:
                    items_data[item.name]["max_price"] = item.unit_price
        
        # Calculate average prices
        items_list = []
        for item_name_key, item_info in items_data.items():
            avg_price = round(item_info["total_price"] / item_info["total_quantity"], 2) if item_info["total_quantity"] > 0 else 0
            items_list.append({
                "name": item_info["name"],
                "unit": item_info["unit"],
                "total_quantity": item_info["total_quantity"],
                "total_price": round(item_info["total_price"], 2),
                "order_count": item_info["order_count"],
                "avg_price": avg_price,
                "min_price": item_info["min_price"] or 0,
                "max_price": item_info["max_price"] or 0,
                "price_history": item_info["prices"][-5:]  # Last 5 price entries
            })
        
        # Sort items by total quantity
        items_list.sort(key=lambda x: x["total_quantity"], reverse=True)
        
        # Calculate on-time rate
        delivered_count = on_time_deliveries + late_deliveries
        on_time_rate = round((on_time_deliveries / delivered_count * 100), 1) if delivered_count > 0 else 0
        
        performance_data.append({
            "supplier_id": supplier.id,
            "supplier_name": supplier.name,
            "contact_person": supplier.contact_person,
            "phone": supplier.phone,
            "email": supplier.email,
            "total_orders": total_orders,
            "completed_orders": completed_orders,
            "approved_orders": approved_orders,
            "total_amount": round(total_amount, 2),
            "completion_rate": round((completed_orders / total_orders * 100), 1) if total_orders > 0 else 0,
            "avg_order_value": round(total_amount / approved_orders, 2) if approved_orders > 0 else 0,
            # Delivery performance
            "on_time_deliveries": on_time_deliveries,
            "late_deliveries": late_deliveries,
            "pending_late": pending_late,
            "on_time_rate": on_time_rate,
            # Items data
            "items": items_list[:20],  # Top 20 items
            "total_items": len(items_list)
        })
    
    # Sort by total amount
    performance_data.sort(key=lambda x: x["total_amount"], reverse=True)
    
    return {
        "suppliers": performance_data,
        "summary": {
            "total_suppliers": len(performance_data),
            "total_orders": sum(s["total_orders"] for s in performance_data),
            "total_spending": round(sum(s["total_amount"] for s in performance_data), 2),
            "avg_on_time_rate": round(sum(s["on_time_rate"] for s in performance_data) / len(performance_data), 1) if performance_data else 0
        },
        "filters_applied": {
            "supplier_id": supplier_id,
            "project_id": project_id,
            "start_date": start_date,
            "end_date": end_date,
            "item_name": item_name
        }
    }


# ==================== PRICE VARIANCE REPORT ====================

@pg_settings_router.get("/reports/advanced/price-variance")
async def get_price_variance_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    item_name: Optional[str] = None,
    period: str = "monthly",  # monthly, quarterly, yearly
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """
    تقرير اختلاف الأسعار - تحليل زمني
    يعرض الأصناف التي تغيّر سعرها عبر الزمن
    """
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.GENERAL_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    # Base query for order items
    query = select(
        PurchaseOrderItem,
        PurchaseOrder
    ).join(
        PurchaseOrder, PurchaseOrderItem.order_id == PurchaseOrder.id
    ).where(
        PurchaseOrder.status.in_(['approved', 'printed', 'shipped', 'delivered', 'partially_delivered'])
    )
    
    # Apply date filters
    if start_date:
        try:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            query = query.where(PurchaseOrder.created_at >= start)
        except:
            pass
    
    if end_date:
        try:
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query = query.where(PurchaseOrder.created_at <= end)
        except:
            pass
    
    # Filter by item name
    if item_name:
        query = query.where(PurchaseOrderItem.name.ilike(f"%{item_name}%"))
    
    query = query.order_by(PurchaseOrder.created_at)
    result = await session.execute(query)
    rows = result.all()
    
    # Collect price data by item
    item_prices = {}  # {item_name: [{date, price, supplier, order_number, quantity}]}
    
    for item, order in rows:
        item_key = item.name.strip().lower()
        
        if item_key not in item_prices:
            item_prices[item_key] = {
                "name": item.name,
                "unit": item.unit,
                "prices": []
            }
        
        item_prices[item_key]["prices"].append({
            "date": order.created_at.strftime("%Y-%m-%d") if order.created_at else None,
            "price": item.unit_price,
            "supplier": order.supplier_name,
            "order_number": order.order_number,
            "quantity": item.quantity
        })
    
    # Analyze price changes
    variance_report = []
    increased_items = []
    decreased_items = []
    
    for item_key, data in item_prices.items():
        if len(data["prices"]) < 2:
            continue  # Need at least 2 prices to compare
        
        prices = data["prices"]
        first_price = prices[0]["price"]
        last_price = prices[-1]["price"]
        min_price = min(p["price"] for p in prices)
        max_price = max(p["price"] for p in prices)
        avg_price = sum(p["price"] for p in prices) / len(prices)
        
        price_change = last_price - first_price
        price_change_percent = round((price_change / first_price * 100), 1) if first_price > 0 else 0
        
        variance_data = {
            "name": data["name"],
            "unit": data["unit"],
            "first_price": round(first_price, 2),
            "last_price": round(last_price, 2),
            "min_price": round(min_price, 2),
            "max_price": round(max_price, 2),
            "avg_price": round(avg_price, 2),
            "price_change": round(price_change, 2),
            "price_change_percent": price_change_percent,
            "variance_count": len(prices),
            "price_history": prices[-10:],  # Last 10 entries
            "trend": "increased" if price_change > 0 else "decreased" if price_change < 0 else "stable"
        }
        
        variance_report.append(variance_data)
        
        if price_change > 0:
            increased_items.append(variance_data)
        elif price_change < 0:
            decreased_items.append(variance_data)
    
    # Sort by absolute price change
    variance_report.sort(key=lambda x: abs(x["price_change"]), reverse=True)
    increased_items.sort(key=lambda x: x["price_change"], reverse=True)
    decreased_items.sort(key=lambda x: x["price_change"])
    
    return {
        "items": variance_report[:50],  # Top 50 items with price changes
        "summary": {
            "total_items_analyzed": len(item_prices),
            "items_with_changes": len(variance_report),
            "increased_items": len(increased_items),
            "decreased_items": len(decreased_items),
            "stable_items": len(item_prices) - len(variance_report)
        },
        "increased": increased_items[:10],  # Top 10 increased
        "decreased": decreased_items[:10],  # Top 10 decreased
        "filters": {
            "start_date": start_date,
            "end_date": end_date,
            "item_name": item_name,
            "period": period
        }
    }


@pg_settings_router.get("/reports/advanced/price-variance/export")
async def export_price_variance_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "excel",
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تصدير تقرير اختلاف الأسعار"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.GENERAL_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    # Get data
    report_data = await get_price_variance_report(
        start_date=start_date,
        end_date=end_date,
        current_user=current_user,
        session=session
    )
    
    if format == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise HTTPException(status_code=500, detail="مكتبة Excel غير متوفرة")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير اختلاف الأسعار"
        ws.sheet_view.rightToLeft = True
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="9333EA", end_color="9333EA", fill_type="solid")
        increase_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        decrease_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = "تقرير اختلاف الأسعار (التحليل الزمني)"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['الصنف', 'الوحدة', 'أول سعر', 'آخر سعر', 'أقل سعر', 'أعلى سعر', 'التغير', 'نسبة التغير %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Data
        for row_num, item in enumerate(report_data["items"], 4):
            ws.cell(row=row_num, column=1, value=item["name"]).border = thin_border
            ws.cell(row=row_num, column=2, value=item["unit"]).border = thin_border
            ws.cell(row=row_num, column=3, value=item["first_price"]).border = thin_border
            ws.cell(row=row_num, column=4, value=item["last_price"]).border = thin_border
            ws.cell(row=row_num, column=5, value=item["min_price"]).border = thin_border
            ws.cell(row=row_num, column=6, value=item["max_price"]).border = thin_border
            
            change_cell = ws.cell(row=row_num, column=7, value=item["price_change"])
            change_cell.border = thin_border
            
            percent_cell = ws.cell(row=row_num, column=8, value=f"{item['price_change_percent']}%")
            percent_cell.border = thin_border
            
            # Color code based on trend
            if item["price_change"] > 0:
                for c in range(1, 9):
                    ws.cell(row=row_num, column=c).fill = increase_fill
            elif item["price_change"] < 0:
                for c in range(1, 9):
                    ws.cell(row=row_num, column=c).fill = decrease_fill
        
        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 12
        for c in 'CDEFGH':
            ws.column_dimensions[c].width = 14
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=price_variance_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    
    return report_data


# ==================== EXPORT ENDPOINTS ====================
from fastapi.responses import StreamingResponse
import io

@pg_settings_router.get("/reports/advanced/summary/export")
async def export_summary_report(
    project_id: Optional[str] = None,
    engineer_id: Optional[str] = None,
    supervisor_id: Optional[str] = None,
    supplier_id: Optional[str] = None,
    format: str = "excel",
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تصدير الملخص التنفيذي إلى Excel"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.GENERAL_MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Get data with filters
        filters = []
        if project_id:
            filters.append(PurchaseOrder.project_id == project_id)
        if supplier_id:
            filters.append(PurchaseOrder.supplier_id == supplier_id)
        
        orders_query = select(PurchaseOrder).where(
            PurchaseOrder.status.in_(["approved", "printed", "shipped", "delivered"])
        )
        if filters:
            orders_query = orders_query.where(and_(*filters))
        
        orders_result = await session.execute(orders_query)
        orders = orders_result.scalars().all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الملخص التنفيذي"
        ws.sheet_view.rightToLeft = True
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "تقرير الملخص التنفيذي"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Summary
        ws['A3'] = "إجمالي أوامر الشراء:"
        ws['B3'] = len(orders)
        ws['A4'] = "إجمالي المصروفات:"
        ws['B4'] = sum(o.total_amount or 0 for o in orders)
        
        # Orders by project
        ws['A7'] = "المشروع"
        ws['B7'] = "عدد الطلبات"
        ws['C7'] = "المبلغ"
        for cell in ['A7', 'B7', 'C7']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
        
        by_project = {}
        for order in orders:
            pname = order.project_name or "غير محدد"
            if pname not in by_project:
                by_project[pname] = {"count": 0, "amount": 0}
            by_project[pname]["count"] += 1
            by_project[pname]["amount"] += order.total_amount or 0
        
        row = 8
        for pname, data in by_project.items():
            ws[f'A{row}'] = pname
            ws[f'B{row}'] = data["count"]
            ws[f'C{row}'] = data["amount"]
            row += 1
        
        # Set column widths manually
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=summary_report.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل في تصدير التقرير: {str(e)}")


@pg_settings_router.get("/reports/advanced/approval-analytics/export")
async def export_approval_report(
    project_id: Optional[str] = None,
    engineer_id: Optional[str] = None,
    supervisor_id: Optional[str] = None,
    format: str = "excel",
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تصدير تحليل الاعتمادات إلى Excel"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.GENERAL_MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Get data with filters
        filters = []
        if project_id:
            filters.append(MaterialRequest.project_id == project_id)
        if engineer_id:
            filters.append(MaterialRequest.engineer_id == engineer_id)
        if supervisor_id:
            filters.append(MaterialRequest.supervisor_id == supervisor_id)
        
        query = select(MaterialRequest)
        if filters:
            query = query.where(and_(*filters))
        
        result = await session.execute(query)
        requests = result.scalars().all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تحليل الاعتمادات"
        ws.sheet_view.rightToLeft = True
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "تقرير تحليل الاعتمادات"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Summary
        total = len(requests)
        approved = len([r for r in requests if r.status in ["approved", "po_issued"]])
        rejected = len([r for r in requests if r.status in ["rejected", "rejected_engineer"]])
        
        ws['A3'] = "إجمالي الطلبات:"
        ws['B3'] = total
        ws['A4'] = "معتمدة:"
        ws['B4'] = approved
        ws['A5'] = "مرفوضة:"
        ws['B5'] = rejected
        
        # By engineer
        ws['A8'] = "المهندس"
        ws['B8'] = "الإجمالي"
        ws['C8'] = "معتمدة"
        ws['D8'] = "مرفوضة"
        for cell in ['A8', 'B8', 'C8', 'D8']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
        
        by_engineer = {}
        for req in requests:
            eng = req.engineer_name or "غير محدد"
            if eng not in by_engineer:
                by_engineer[eng] = {"total": 0, "approved": 0, "rejected": 0}
            by_engineer[eng]["total"] += 1
            if req.status in ["approved", "po_issued"]:
                by_engineer[eng]["approved"] += 1
            elif req.status in ["rejected", "rejected_engineer"]:
                by_engineer[eng]["rejected"] += 1
        
        row = 9
        for eng, data in by_engineer.items():
            ws[f'A{row}'] = eng
            ws[f'B{row}'] = data["total"]
            ws[f'C{row}'] = data["approved"]
            ws[f'D{row}'] = data["rejected"]
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=approval_report.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل في تصدير التقرير: {str(e)}")


@pg_settings_router.get("/reports/advanced/supplier-performance/export")
async def export_supplier_report(
    supplier_id: Optional[str] = None,
    format: str = "excel",
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تصدير أداء الموردين إلى Excel"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.GENERAL_MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Get suppliers
        query = select(Supplier)
        if supplier_id:
            query = query.where(Supplier.id == supplier_id)
        result = await session.execute(query)
        suppliers = result.scalars().all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "أداء الموردين"
        ws.sheet_view.rightToLeft = True
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = "تقرير أداء الموردين"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Headers
        headers = ["المورد", "جهة الاتصال", "الهاتف", "عدد الطلبات", "المكتملة", "إجمالي المشتريات", "متوسط الطلب"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row = 4
        total_orders = 0
        total_spending = 0
        
        for supplier in suppliers:
            # Get orders for this supplier
            orders_result = await session.execute(
                select(PurchaseOrder).where(PurchaseOrder.supplier_id == supplier.id)
            )
            orders = orders_result.scalars().all()
            
            order_count = len(orders)
            completed = len([o for o in orders if o.status == "delivered"])
            amount = sum(o.total_amount or 0 for o in orders if o.status in ["approved", "printed", "shipped", "delivered"])
            avg = amount / order_count if order_count > 0 else 0
            
            ws.cell(row=row, column=1, value=supplier.name)
            ws.cell(row=row, column=2, value=supplier.contact_person or "-")
            ws.cell(row=row, column=3, value=supplier.phone or "-")
            ws.cell(row=row, column=4, value=order_count)
            ws.cell(row=row, column=5, value=completed)
            ws.cell(row=row, column=6, value=amount)
            ws.cell(row=row, column=7, value=round(avg, 2))
            
            total_orders += order_count
            total_spending += amount
            row += 1
        
        # Summary row
        ws.cell(row=row+1, column=1, value="الإجمالي")
        ws.cell(row=row+1, column=1).font = Font(bold=True)
        ws.cell(row=row+1, column=4, value=total_orders)
        ws.cell(row=row+1, column=6, value=total_spending)
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=supplier_report.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل في تصدير التقرير: {str(e)}")


# ==================== AUDIT LOG ROUTES ====================

@pg_settings_router.get("/audit-logs")
async def get_audit_logs(
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get audit logs - system admin and procurement manager only"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    query = select(AuditLog)
    
    if entity_type:
        query = query.where(AuditLog.entity_type == entity_type)
    if entity_id:
        query = query.where(AuditLog.entity_id == entity_id)
    
    query = query.order_by(desc(AuditLog.timestamp)).limit(limit)
    
    result = await session.execute(query)
    logs = result.scalars().all()
    
    return [
        {
            "id": log.id,
            "entity_type": log.entity_type,
            "entity_id": log.entity_id,
            "action": log.action,
            "description": log.description,
            "user_name": log.user_name,
            "user_role": log.user_role,
            "changes": json.loads(log.changes) if log.changes else None,
            "timestamp": log.timestamp.isoformat() if log.timestamp else None
        }
        for log in logs
    ]
