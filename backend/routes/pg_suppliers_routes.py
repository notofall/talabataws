"""
PostgreSQL Suppliers Routes - Supplier Management
Migrated from MongoDB to PostgreSQL
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timedelta
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func, desc, and_, or_
import uuid
import io

from database import get_postgres_session, Supplier, User, PurchaseOrder, PurchaseOrderItem, PriceCatalogItem

# Create router
pg_suppliers_router = APIRouter(prefix="/api/pg", tags=["PostgreSQL Suppliers"])

# Import auth dependency
from routes.pg_auth_routes import get_current_user_pg, UserRole


# ==================== PYDANTIC MODELS ====================

class SupplierCreate(BaseModel):
    name: str
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None


class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None


class SupplierResponse(BaseModel):
    id: str
    name: str
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    created_at: str


# ==================== SUPPLIERS ROUTES ====================

@pg_suppliers_router.post("/suppliers")
async def create_supplier(
    supplier_data: SupplierCreate,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Create a new supplier - procurement manager only"""
    if current_user.role != UserRole.PROCUREMENT_MANAGER:
        raise HTTPException(status_code=403, detail="فقط مدير المشتريات يمكنه إدارة الموردين")
    
    supplier_id = str(uuid.uuid4())
    now = datetime.utcnow()
    
    new_supplier = Supplier(
        id=supplier_id,
        name=supplier_data.name,
        contact_person=supplier_data.contact_person,
        phone=supplier_data.phone,
        email=supplier_data.email,
        address=supplier_data.address,
        notes=supplier_data.notes,
        created_at=now
    )
    
    session.add(new_supplier)
    await session.commit()
    
    return {
        "id": supplier_id,
        "name": supplier_data.name,
        "contact_person": supplier_data.contact_person,
        "phone": supplier_data.phone,
        "email": supplier_data.email,
        "address": supplier_data.address,
        "notes": supplier_data.notes,
        "created_at": now.isoformat()
    }


@pg_suppliers_router.get("/suppliers")
async def get_suppliers(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get all suppliers"""
    result = await session.execute(
        select(Supplier).order_by(desc(Supplier.created_at))
    )
    suppliers = result.scalars().all()
    
    return [
        {
            "id": s.id,
            "name": s.name,
            "contact_person": s.contact_person,
            "phone": s.phone,
            "email": s.email,
            "address": s.address,
            "notes": s.notes,
            "created_at": s.created_at.isoformat() if s.created_at else None
        }
        for s in suppliers
    ]


@pg_suppliers_router.get("/suppliers/{supplier_id}")
async def get_supplier(
    supplier_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get a single supplier by ID"""
    result = await session.execute(select(Supplier).where(Supplier.id == supplier_id))
    supplier = result.scalar_one_or_none()
    
    if not supplier:
        raise HTTPException(status_code=404, detail="المورد غير موجود")
    
    # Get supplier stats
    order_result = await session.execute(
        select(
            func.count().label('count'),
            func.coalesce(func.sum(PurchaseOrder.total_amount), 0).label('total')
        ).select_from(PurchaseOrder).where(PurchaseOrder.supplier_id == supplier_id)
    )
    stats = order_result.first()
    
    return {
        "id": supplier.id,
        "name": supplier.name,
        "contact_person": supplier.contact_person,
        "phone": supplier.phone,
        "email": supplier.email,
        "address": supplier.address,
        "notes": supplier.notes,
        "created_at": supplier.created_at.isoformat() if supplier.created_at else None,
        "total_orders": stats.count if stats else 0,
        "total_value": float(stats.total) if stats else 0
    }


@pg_suppliers_router.put("/suppliers/{supplier_id}")
async def update_supplier(
    supplier_id: str,
    update_data: SupplierUpdate,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Update a supplier - procurement manager only"""
    if current_user.role != UserRole.PROCUREMENT_MANAGER:
        raise HTTPException(status_code=403, detail="فقط مدير المشتريات يمكنه تعديل الموردين")
    
    result = await session.execute(select(Supplier).where(Supplier.id == supplier_id))
    supplier = result.scalar_one_or_none()
    
    if not supplier:
        raise HTTPException(status_code=404, detail="المورد غير موجود")
    
    if update_data.name is not None:
        supplier.name = update_data.name
    if update_data.contact_person is not None:
        supplier.contact_person = update_data.contact_person
    if update_data.phone is not None:
        supplier.phone = update_data.phone
    if update_data.email is not None:
        supplier.email = update_data.email
    if update_data.address is not None:
        supplier.address = update_data.address
    if update_data.notes is not None:
        supplier.notes = update_data.notes
    
    await session.commit()
    
    return {"message": "تم تحديث المورد بنجاح"}


@pg_suppliers_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(
    supplier_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Delete a supplier - procurement manager only"""
    if current_user.role != UserRole.PROCUREMENT_MANAGER:
        raise HTTPException(status_code=403, detail="فقط مدير المشتريات يمكنه حذف الموردين")
    
    # Check if supplier has orders
    order_result = await session.execute(
        select(func.count()).select_from(PurchaseOrder).where(PurchaseOrder.supplier_id == supplier_id)
    )
    order_count = order_result.scalar() or 0
    
    if order_count > 0:
        raise HTTPException(status_code=400, detail=f"لا يمكن حذف المورد لوجود {order_count} أوامر شراء مرتبطة به")
    
    result = await session.execute(select(Supplier).where(Supplier.id == supplier_id))
    supplier = result.scalar_one_or_none()
    
    if not supplier:
        raise HTTPException(status_code=404, detail="المورد غير موجود")
    
    await session.delete(supplier)
    await session.commit()
    
    return {"message": "تم حذف المورد بنجاح"}


# ==================== SUPPLIER PERFORMANCE REPORT ====================

@pg_suppliers_router.get("/suppliers/performance/report")
async def get_supplier_performance_report(
    supplier_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    item_name: Optional[str] = None,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get supplier performance report with filters"""
    
    # Base query for orders
    query = select(PurchaseOrder).where(
        PurchaseOrder.status.in_(['approved', 'printed', 'shipped', 'delivered', 'partially_delivered'])
    )
    
    # Apply supplier filter
    if supplier_id:
        query = query.where(PurchaseOrder.supplier_id == supplier_id)
    
    # Apply date filters
    if start_date:
        try:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            query = query.where(PurchaseOrder.created_at >= start)
        except:
            pass
    
    if end_date:
        try:
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query = query.where(PurchaseOrder.created_at <= end)
        except:
            pass
    
    result = await session.execute(query.order_by(desc(PurchaseOrder.created_at)))
    orders = result.scalars().all()
    
    # Get all suppliers for the report
    suppliers_query = select(Supplier)
    if supplier_id:
        suppliers_query = suppliers_query.where(Supplier.id == supplier_id)
    
    suppliers_result = await session.execute(suppliers_query)
    suppliers = suppliers_result.scalars().all()
    
    # Build supplier performance data
    supplier_data = {}
    for supplier in suppliers:
        supplier_data[supplier.id] = {
            "id": supplier.id,
            "name": supplier.name,
            "contact_person": supplier.contact_person,
            "phone": supplier.phone,
            "email": supplier.email,
            "total_orders": 0,
            "total_amount": 0,
            "delivered_orders": 0,
            "on_time_deliveries": 0,
            "late_deliveries": 0,
            "items": {},  # {item_name: {quantity, total_price, prices: []}}
            "orders_list": []
        }
    
    # Process orders
    for order in orders:
        if order.supplier_id not in supplier_data:
            continue
        
        data = supplier_data[order.supplier_id]
        data["total_orders"] += 1
        data["total_amount"] += order.total_amount or 0
        
        # Check delivery status
        if order.status in ['delivered', 'partially_delivered']:
            data["delivered_orders"] += 1
            
            # Check if on time (delivered_at vs expected_delivery_date)
            if order.expected_delivery_date and order.delivered_at:
                try:
                    expected = datetime.strptime(order.expected_delivery_date, "%Y-%m-%d")
                    if order.delivered_at <= expected:
                        data["on_time_deliveries"] += 1
                    else:
                        data["late_deliveries"] += 1
                except:
                    pass
        
        # Get order items
        items_result = await session.execute(
            select(PurchaseOrderItem).where(PurchaseOrderItem.order_id == order.id)
        )
        items = items_result.scalars().all()
        
        for item in items:
            # Filter by item name if specified
            if item_name and item_name.lower() not in item.name.lower():
                continue
            
            if item.name not in data["items"]:
                data["items"][item.name] = {
                    "name": item.name,
                    "unit": item.unit,
                    "total_quantity": 0,
                    "total_price": 0,
                    "prices": [],
                    "order_count": 0
                }
            
            data["items"][item.name]["total_quantity"] += item.quantity
            data["items"][item.name]["total_price"] += item.total_price or 0
            data["items"][item.name]["prices"].append({
                "unit_price": item.unit_price,
                "quantity": item.quantity,
                "date": order.created_at.isoformat() if order.created_at else None
            })
            data["items"][item.name]["order_count"] += 1
        
        # Add to orders list
        data["orders_list"].append({
            "order_number": order.order_number,
            "project_name": order.project_name,
            "total_amount": order.total_amount,
            "status": order.status,
            "created_at": order.created_at.isoformat() if order.created_at else None,
            "expected_delivery": order.expected_delivery_date,
            "delivered_at": order.delivered_at.isoformat() if order.delivered_at else None
        })
    
    # Calculate averages and format response
    report = []
    for sup_id, data in supplier_data.items():
        # Calculate on-time rate
        on_time_rate = 0
        if data["delivered_orders"] > 0:
            on_time_rate = round((data["on_time_deliveries"] / data["delivered_orders"]) * 100, 1)
        
        # Convert items dict to list
        items_list = list(data["items"].values())
        for item in items_list:
            if item["prices"]:
                item["avg_price"] = round(item["total_price"] / item["total_quantity"], 2) if item["total_quantity"] > 0 else 0
                item["min_price"] = min(p["unit_price"] for p in item["prices"])
                item["max_price"] = max(p["unit_price"] for p in item["prices"])
        
        report.append({
            "supplier": {
                "id": data["id"],
                "name": data["name"],
                "contact_person": data["contact_person"],
                "phone": data["phone"],
                "email": data["email"]
            },
            "performance": {
                "total_orders": data["total_orders"],
                "total_amount": round(data["total_amount"], 2),
                "delivered_orders": data["delivered_orders"],
                "on_time_deliveries": data["on_time_deliveries"],
                "late_deliveries": data["late_deliveries"],
                "on_time_rate": on_time_rate
            },
            "items": items_list,
            "recent_orders": data["orders_list"][:10]  # Last 10 orders
        })
    
    # Sort by total orders descending
    report.sort(key=lambda x: x["performance"]["total_orders"], reverse=True)
    
    return {
        "report": report,
        "summary": {
            "total_suppliers": len(report),
            "total_orders": sum(r["performance"]["total_orders"] for r in report),
            "total_amount": round(sum(r["performance"]["total_amount"] for r in report), 2)
        },
        "filters": {
            "supplier_id": supplier_id,
            "start_date": start_date,
            "end_date": end_date,
            "item_name": item_name
        }
    }


@pg_suppliers_router.get("/suppliers/performance/export")
async def export_supplier_performance(
    supplier_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "excel",
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Export supplier performance report to Excel/PDF"""
    # Get report data
    report_data = await get_supplier_performance_report(
        supplier_id=supplier_id,
        start_date=start_date,
        end_date=end_date,
        current_user=current_user,
        session=session
    )
    
    if format == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise HTTPException(status_code=500, detail="مكتبة Excel غير متوفرة")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير أداء الموردين"
        ws.sheet_view.rightToLeft = True
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['المورد', 'إجمالي الأوامر', 'إجمالي المبلغ', 'الأوامر المسلمة', 'في الوقت', 'متأخرة', 'نسبة الالتزام %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Data
        for row_num, r in enumerate(report_data["report"], 2):
            ws.cell(row=row_num, column=1, value=r["supplier"]["name"]).border = thin_border
            ws.cell(row=row_num, column=2, value=r["performance"]["total_orders"]).border = thin_border
            ws.cell(row=row_num, column=3, value=r["performance"]["total_amount"]).border = thin_border
            ws.cell(row=row_num, column=4, value=r["performance"]["delivered_orders"]).border = thin_border
            ws.cell(row=row_num, column=5, value=r["performance"]["on_time_deliveries"]).border = thin_border
            ws.cell(row=row_num, column=6, value=r["performance"]["late_deliveries"]).border = thin_border
            ws.cell(row=row_num, column=7, value=f"{r['performance']['on_time_rate']}%").border = thin_border
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=supplier_performance_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    
    return report_data
