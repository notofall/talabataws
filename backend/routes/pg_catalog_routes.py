"""
PostgreSQL Price Catalog Routes
Routes for managing price catalog and item aliases
"""
from fastapi import APIRouter, HTTPException, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func, or_, desc
import uuid
import json
import io
import csv

from database import get_postgres_session, User, PriceCatalogItem, ItemAlias, BudgetCategory, Supplier, PurchaseOrder, PurchaseOrderItem
from routes.pg_auth_routes import get_current_user_pg, UserRole

# Create router
pg_catalog_router = APIRouter(prefix="/api/pg", tags=["PostgreSQL Catalog"])


# ==================== PYDANTIC MODELS ====================

class CatalogItemCreate(BaseModel):
    item_code: Optional[str] = None  # كود الصنف
    name: str
    description: Optional[str] = None
    unit: str = "قطعة"
    supplier_id: Optional[str] = None
    supplier_name: Optional[str] = None
    price: float
    currency: str = "SAR"
    validity_until: Optional[str] = None
    category_id: Optional[str] = None
    category_name: Optional[str] = None


class CatalogItemUpdate(BaseModel):
    item_code: Optional[str] = None  # كود الصنف
    name: Optional[str] = None
    description: Optional[str] = None
    unit: Optional[str] = None
    supplier_id: Optional[str] = None
    supplier_name: Optional[str] = None
    price: Optional[float] = None
    currency: Optional[str] = None
    validity_until: Optional[str] = None
    category_id: Optional[str] = None
    category_name: Optional[str] = None
    is_active: Optional[bool] = None


class AliasCreate(BaseModel):
    alias_name: str
    catalog_item_id: str


class AliasUpdate(BaseModel):
    alias_name: Optional[str] = None
    catalog_item_id: Optional[str] = None


# ==================== PRICE CATALOG ROUTES ====================

@pg_catalog_router.get("/price-catalog")
async def get_price_catalog(
    search: str = "",
    category_id: Optional[str] = None,
    supplier_id: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get price catalog items with pagination and search"""
    query = select(PriceCatalogItem).where(PriceCatalogItem.is_active == True)
    
    # Apply search filter - include item_code in search
    if search:
        query = query.where(
            or_(
                PriceCatalogItem.name.ilike(f"%{search}%"),
                PriceCatalogItem.description.ilike(f"%{search}%"),
                PriceCatalogItem.item_code.ilike(f"%{search}%")
            )
        )
    
    # Apply category filter
    if category_id:
        query = query.where(PriceCatalogItem.category_id == category_id)
    
    # Apply supplier filter
    if supplier_id:
        query = query.where(PriceCatalogItem.supplier_id == supplier_id)
    
    # Count total
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await session.execute(count_query)
    total = total_result.scalar()
    
    # Apply pagination
    query = query.order_by(PriceCatalogItem.item_code.asc().nullslast(), PriceCatalogItem.name).offset((page - 1) * page_size).limit(page_size)
    
    result = await session.execute(query)
    items = result.scalars().all()
    
    return {
        "items": [
            {
                "id": item.id,
                "item_code": item.item_code,
                "name": item.name,
                "description": item.description,
                "unit": item.unit,
                "supplier_id": item.supplier_id,
                "supplier_name": item.supplier_name,
                "price": item.price,
                "currency": item.currency,
                "validity_until": item.validity_until,
                "category_id": item.category_id,
                "category_name": item.category_name,
                "is_active": item.is_active,
                "created_by": item.created_by,
                "created_by_name": item.created_by_name,
                "created_at": item.created_at.isoformat() if item.created_at else None
            }
            for item in items
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@pg_catalog_router.post("/price-catalog")
async def create_catalog_item(
    item_data: CatalogItemCreate,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Create a new catalog item - procurement manager only"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بإضافة عناصر للكتالوج")
    
    # Check if item_code already exists
    if item_data.item_code:
        existing = await session.execute(
            select(PriceCatalogItem).where(PriceCatalogItem.item_code == item_data.item_code)
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail=f"كود الصنف '{item_data.item_code}' موجود بالفعل")
    
    # Validate and get category name if category_id provided
    category_name = item_data.category_name
    valid_category_id = None
    if item_data.category_id:
        cat_result = await session.execute(
            select(BudgetCategory).where(BudgetCategory.id == item_data.category_id)
        )
        category = cat_result.scalar_one_or_none()
        if category:
            valid_category_id = item_data.category_id
            category_name = category.name
    
    # Validate and get supplier name if supplier_id provided
    supplier_name = item_data.supplier_name
    valid_supplier_id = None
    if item_data.supplier_id:
        sup_result = await session.execute(
            select(Supplier).where(Supplier.id == item_data.supplier_id)
        )
        supplier = sup_result.scalar_one_or_none()
        if supplier:
            valid_supplier_id = item_data.supplier_id
            supplier_name = supplier.name
    
    new_item = PriceCatalogItem(
        id=str(uuid.uuid4()),
        item_code=item_data.item_code,
        name=item_data.name,
        description=item_data.description,
        unit=item_data.unit,
        supplier_id=valid_supplier_id,
        supplier_name=supplier_name,
        price=item_data.price,
        currency=item_data.currency,
        validity_until=item_data.validity_until,
        category_id=valid_category_id,
        category_name=category_name,
        is_active=True,
        created_by=current_user.id,
        created_by_name=current_user.name
    )
    
    session.add(new_item)
    await session.commit()
    await session.refresh(new_item)
    
    return {
        "id": new_item.id,
        "item_code": new_item.item_code,
        "name": new_item.name,
        "price": new_item.price,
        "message": "تم إضافة العنصر للكتالوج بنجاح"
    }


@pg_catalog_router.put("/price-catalog/{item_id}")
async def update_catalog_item(
    item_id: str,
    item_data: CatalogItemUpdate,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Update a catalog item"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل الكتالوج")
    
    result = await session.execute(
        select(PriceCatalogItem).where(PriceCatalogItem.id == item_id)
    )
    item = result.scalar_one_or_none()
    
    if not item:
        raise HTTPException(status_code=404, detail="العنصر غير موجود")
    
    # Check if item_code is being changed and if it already exists
    if item_data.item_code and item_data.item_code != item.item_code:
        existing = await session.execute(
            select(PriceCatalogItem).where(
                PriceCatalogItem.item_code == item_data.item_code,
                PriceCatalogItem.id != item_id
            )
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail=f"كود الصنف '{item_data.item_code}' موجود بالفعل")
    
    # Update fields
    update_data = item_data.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(item, key, value)
    
    item.updated_at = datetime.utcnow()
    
    await session.commit()
    
    return {"message": "تم تحديث العنصر بنجاح"}


@pg_catalog_router.delete("/price-catalog/{item_id}")
async def delete_catalog_item(
    item_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Delete (deactivate) a catalog item"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بحذف عناصر الكتالوج")
    
    result = await session.execute(
        select(PriceCatalogItem).where(PriceCatalogItem.id == item_id)
    )
    item = result.scalar_one_or_none()
    
    if not item:
        raise HTTPException(status_code=404, detail="العنصر غير موجود")
    
    # Soft delete
    item.is_active = False
    item.updated_at = datetime.utcnow()
    
    await session.commit()
    
    return {"message": "تم حذف العنصر بنجاح"}


@pg_catalog_router.get("/price-catalog/export")
async def export_catalog(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Export catalog to CSV"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بتصدير الكتالوج")
    
    result = await session.execute(
        select(PriceCatalogItem).where(PriceCatalogItem.is_active == True).order_by(PriceCatalogItem.name)
    )
    items = result.scalars().all()
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(['الاسم', 'الوصف', 'الوحدة', 'السعر', 'العملة', 'المورد', 'التصنيف', 'صالح حتى'])
    
    # Data
    for item in items:
        writer.writerow([
            item.name,
            item.description or '',
            item.unit,
            item.price,
            item.currency,
            item.supplier_name or '',
            item.category_name or '',
            item.validity_until or ''
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=price_catalog_{datetime.now().strftime('%Y%m%d')}.csv"}
    )


@pg_catalog_router.get("/price-catalog/export/excel")
async def export_catalog_excel(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Export catalog to Excel"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بتصدير الكتالوج")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة Excel غير متوفرة")
    
    result = await session.execute(
        select(PriceCatalogItem).where(PriceCatalogItem.is_active == True).order_by(PriceCatalogItem.name)
    )
    items = result.scalars().all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "كتالوج الأسعار"
    ws.sheet_view.rightToLeft = True
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['الاسم', 'الوصف', 'الوحدة', 'السعر', 'العملة', 'المورد', 'التصنيف', 'صالح حتى']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data
    for row_num, item in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=item.name).border = thin_border
        ws.cell(row=row_num, column=2, value=item.description or '').border = thin_border
        ws.cell(row=row_num, column=3, value=item.unit).border = thin_border
        ws.cell(row=row_num, column=4, value=item.price).border = thin_border
        ws.cell(row=row_num, column=5, value=item.currency).border = thin_border
        ws.cell(row=row_num, column=6, value=item.supplier_name or '').border = thin_border
        ws.cell(row=row_num, column=7, value=item.category_name or '').border = thin_border
        ws.cell(row=row_num, column=8, value=item.validity_until or '').border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=price_catalog_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


@pg_catalog_router.get("/price-catalog/template")
async def get_catalog_template(
    current_user: User = Depends(get_current_user_pg)
):
    """Download Excel template for catalog import"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة Excel غير متوفرة")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "نموذج الكتالوج"
    ws.sheet_view.rightToLeft = True
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers - must match import order
    headers = ['كود الصنف', 'اسم الصنف', 'الوصف', 'الوحدة', 'السعر', 'العملة', 'اسم المورد', 'التصنيف', 'صالح حتى']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Example row
    example_data = ['ITM-001', 'حديد تسليح 12مم', 'حديد تسليح قطر 12مم', 'طن', '3500', 'SAR', 'شركة الحديد', 'مواد بناء', '2025-12-31']
    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=catalog_template.xlsx"}
    )


@pg_catalog_router.post("/price-catalog/import")
async def import_catalog(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Import catalog from CSV or Excel"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك باستيراد الكتالوج")
    
    filename = file.filename.lower()
    content = await file.read()
    
    imported = 0
    updated = 0
    errors = []
    
    # Handle Excel files
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or not row[1]:  # Skip empty rows, check name column
                        continue
                    
                    item_code = str(row[0]).strip() if row[0] else None
                    name = str(row[1]).strip() if row[1] else None
                    description = str(row[2]).strip() if len(row) > 2 and row[2] else None
                    unit = str(row[3]).strip() if len(row) > 3 and row[3] else "قطعة"
                    price = float(row[4]) if len(row) > 4 and row[4] else 0
                    currency = str(row[5]).strip() if len(row) > 5 and row[5] else "SAR"
                    supplier_name = str(row[6]).strip() if len(row) > 6 and row[6] else None
                    category_name = str(row[7]).strip() if len(row) > 7 and row[7] else None
                    validity_until = str(row[8]).strip() if len(row) > 8 and row[8] else None
                    
                    if not name or price <= 0:
                        continue
                    
                    # Check if item with same code exists
                    existing_item = None
                    if item_code:
                        existing_result = await session.execute(
                            select(PriceCatalogItem).where(
                                PriceCatalogItem.name == name,
                                PriceCatalogItem.supplier_name == supplier_name
                            )
                        )
                        existing_item = existing_result.scalar_one_or_none()
                    
                    if existing_item:
                        # Update existing item
                        existing_item.description = description
                        existing_item.unit = unit
                        existing_item.price = price
                        existing_item.currency = currency
                        existing_item.category_name = category_name
                        existing_item.validity_until = validity_until
                        existing_item.updated_at = datetime.utcnow()
                        updated += 1
                    else:
                        # Create new item
                        new_item = PriceCatalogItem(
                            id=str(uuid.uuid4()),
                            name=name,
                            description=description,
                            unit=unit,
                            price=price,
                            currency=currency,
                            supplier_name=supplier_name,
                            category_name=category_name,
                            validity_until=validity_until,
                            is_active=True,
                            created_by=current_user.id,
                            created_by_name=current_user.name
                        )
                        session.add(new_item)
                        imported += 1
                        
                except Exception as e:
                    errors.append(f"خطأ في السطر {row_num}: {str(e)}")
                    
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"فشل في قراءة ملف Excel: {str(e)}")
    
    # Handle CSV files
    elif filename.endswith('.csv'):
        decoded = content.decode('utf-8-sig')
        reader = csv.reader(io.StringIO(decoded))
        
        # Skip header
        next(reader, None)
        
        for row_num, row in enumerate(reader, start=2):
            try:
                if len(row) < 2:
                    continue
                
                item_code = row[0].strip() if len(row) > 0 else None
                name = row[1].strip() if len(row) > 1 else None
                description = row[2].strip() if len(row) > 2 else None
                unit = row[3].strip() if len(row) > 3 else "قطعة"
                price = float(row[4]) if len(row) > 4 and row[4] else 0
                currency = row[5].strip() if len(row) > 5 else "SAR"
                supplier_name = row[6].strip() if len(row) > 6 else None
                category_name = row[7].strip() if len(row) > 7 else None
                validity_until = row[8].strip() if len(row) > 8 else None
                
                if not name or price <= 0:
                    continue
                
                new_item = PriceCatalogItem(
                    id=str(uuid.uuid4()),
                    name=name,
                    description=description,
                    unit=unit,
                    price=price,
                    currency=currency,
                    supplier_name=supplier_name,
                    category_name=category_name,
                    validity_until=validity_until,
                    is_active=True,
                    created_by=current_user.id,
                    created_by_name=current_user.name
                )
                
                session.add(new_item)
                imported += 1
                
            except Exception as e:
                errors.append(f"خطأ في السطر {row_num}: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="يجب أن يكون الملف بصيغة CSV أو Excel")
    
    await session.commit()
    
    return {
        "message": f"تم استيراد {imported} عنصر جديد وتحديث {updated} عنصر",
        "imported": imported,
        "updated": updated,
        "errors": errors[:10] if errors else []
    }


# ==================== ITEM ALIASES ROUTES ====================

@pg_catalog_router.get("/item-aliases")
async def get_item_aliases(
    search: str = "",
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Get all item aliases"""
    query = select(ItemAlias)
    
    if search:
        query = query.where(ItemAlias.alias_name.ilike(f"%{search}%"))
    
    # Count total
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await session.execute(count_query)
    total = total_result.scalar()
    
    query = query.order_by(ItemAlias.alias_name).offset((page - 1) * page_size).limit(page_size)
    
    result = await session.execute(query)
    aliases = result.scalars().all()
    
    return {
        "items": [
            {
                "id": alias.id,
                "alias_name": alias.alias_name,
                "catalog_item_id": alias.catalog_item_id,
                "catalog_item_name": alias.catalog_item_name,
                "usage_count": alias.usage_count,
                "created_by": alias.created_by,
                "created_by_name": alias.created_by_name,
                "created_at": alias.created_at.isoformat() if alias.created_at else None
            }
            for alias in aliases
        ],
        "total": total,
        "page": page,
        "page_size": page_size
    }


@pg_catalog_router.get("/item-aliases/suggest/{item_name}")
async def suggest_alias(
    item_name: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Suggest catalog item for a given item name"""
    # First check aliases
    alias_result = await session.execute(
        select(ItemAlias).where(ItemAlias.alias_name.ilike(f"%{item_name}%"))
    )
    alias = alias_result.scalar_one_or_none()
    
    if alias:
        # Get the catalog item
        catalog_result = await session.execute(
            select(PriceCatalogItem).where(PriceCatalogItem.id == alias.catalog_item_id)
        )
        catalog_item = catalog_result.scalar_one_or_none()
        
        if catalog_item:
            # Increment usage count
            alias.usage_count += 1
            await session.commit()
            
            return {
                "found": True,
                "source": "alias",
                "catalog_item": {
                    "id": catalog_item.id,
                    "name": catalog_item.name,
                    "price": catalog_item.price,
                    "unit": catalog_item.unit,
                    "supplier_name": catalog_item.supplier_name
                }
            }
    
    # Search in catalog directly
    catalog_result = await session.execute(
        select(PriceCatalogItem).where(
            PriceCatalogItem.is_active == True,
            PriceCatalogItem.name.ilike(f"%{item_name}%")
        ).limit(5)
    )
    items = catalog_result.scalars().all()
    
    if items:
        return {
            "found": True,
            "source": "catalog",
            "suggestions": [
                {
                    "id": item.id,
                    "name": item.name,
                    "price": item.price,
                    "unit": item.unit,
                    "supplier_name": item.supplier_name
                }
                for item in items
            ]
        }
    
    return {"found": False, "suggestions": []}


# ==================== ITEM VALIDATION & BEST PRICE ALERT ====================

class ItemValidationRequest(BaseModel):
    items: List[dict]  # [{name: str, quantity: float, unit: str}]
    supplier_id: Optional[str] = None

class QuickCatalogAdd(BaseModel):
    name: str
    unit: str
    price: float
    currency: str = "SAR"
    supplier_name: Optional[str] = None


@pg_catalog_router.post("/price-catalog/validate-items")
async def validate_items_for_order(
    request: ItemValidationRequest,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """
    التحقق من الأصناف قبل اعتماد أمر الشراء
    - التحقق من وجود كل صنف في الكتالوج
    - إرجاع قائمة الأصناف غير الموجودة
    - إقتراح أصناف مشابهة من الكتالوج
    """
    results = []
    missing_items = []
    
    for item in request.items:
        item_name = item.get("name", "").strip()
        if not item_name:
            continue
        
        # Search in catalog by exact name first
        exact_result = await session.execute(
            select(PriceCatalogItem).where(
                PriceCatalogItem.is_active == True,
                PriceCatalogItem.name == item_name
            )
        )
        exact_match = exact_result.scalars().all()
        
        if exact_match:
            results.append({
                "item_name": item_name,
                "found": True,
                "catalog_items": [
                    {
                        "id": ci.id,
                        "name": ci.name,
                        "price": ci.price,
                        "unit": ci.unit,
                        "currency": ci.currency,
                        "supplier_name": ci.supplier_name
                    }
                    for ci in exact_match
                ]
            })
            continue
        
        # Search for similar items
        similar_result = await session.execute(
            select(PriceCatalogItem).where(
                PriceCatalogItem.is_active == True,
                PriceCatalogItem.name.ilike(f"%{item_name}%")
            ).limit(5)
        )
        similar_items = similar_result.scalars().all()
        
        # Also check aliases
        alias_result = await session.execute(
            select(ItemAlias).where(ItemAlias.alias_name.ilike(f"%{item_name}%"))
        )
        aliases = alias_result.scalars().all()
        
        suggestions = []
        for si in similar_items:
            suggestions.append({
                "id": si.id,
                "name": si.name,
                "price": si.price,
                "unit": si.unit,
                "supplier_name": si.supplier_name
            })
        
        for alias in aliases:
            # Get catalog item for alias
            cat_result = await session.execute(
                select(PriceCatalogItem).where(PriceCatalogItem.id == alias.catalog_item_id)
            )
            cat_item = cat_result.scalar_one_or_none()
            if cat_item and cat_item.id not in [s["id"] for s in suggestions]:
                suggestions.append({
                    "id": cat_item.id,
                    "name": cat_item.name,
                    "price": cat_item.price,
                    "unit": cat_item.unit,
                    "supplier_name": cat_item.supplier_name,
                    "matched_alias": alias.alias_name
                })
        
        results.append({
            "item_name": item_name,
            "found": False,
            "suggestions": suggestions
        })
        missing_items.append({
            "name": item_name,
            "unit": item.get("unit", ""),
            "quantity": item.get("quantity", 0)
        })
    
    return {
        "all_valid": len(missing_items) == 0,
        "total_items": len(request.items),
        "found_items": len(request.items) - len(missing_items),
        "missing_items": len(missing_items),
        "results": results,
        "missing_list": missing_items
    }


@pg_catalog_router.post("/price-catalog/check-best-price")
async def check_best_price(
    item_name: str,
    supplier_id: Optional[str] = None,
    unit_price: float = 0,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """
    تنبيه السعر الأفضل
    - البحث عن نفس الصنف من موردين آخرين بسعر أقل
    - إرجاع قائمة الموردين مع أسعارهم
    """
    # Get supplier name if supplier_id provided
    supplier_name = None
    if supplier_id:
        sup_result = await session.execute(
            select(Supplier).where(Supplier.id == supplier_id)
        )
        supplier = sup_result.scalar_one_or_none()
        if supplier:
            supplier_name = supplier.name
    
    # Search for this item in catalog from other suppliers
    query = select(PriceCatalogItem).where(
        PriceCatalogItem.is_active == True,
        PriceCatalogItem.name.ilike(f"%{item_name}%")
    )
    
    result = await session.execute(query)
    catalog_items = result.scalars().all()
    
    # Also search in historical purchase orders
    po_items_query = select(PurchaseOrderItem).where(
        PurchaseOrderItem.name.ilike(f"%{item_name}%")
    ).order_by(desc(PurchaseOrderItem.id)).limit(20)
    
    po_result = await session.execute(po_items_query)
    po_items = po_result.scalars().all()
    
    # Collect all prices from different suppliers
    supplier_prices = {}  # {supplier_name: {price, source, date}}
    
    for ci in catalog_items:
        sup_name = ci.supplier_name or "غير محدد"
        if sup_name not in supplier_prices or ci.price < supplier_prices[sup_name]["price"]:
            supplier_prices[sup_name] = {
                "price": ci.price,
                "source": "catalog",
                "unit": ci.unit,
                "currency": ci.currency,
                "item_name": ci.name,
                "catalog_item_id": ci.id
            }
    
    # Check historical PO prices
    for poi in po_items:
        # Get order to find supplier
        order_result = await session.execute(
            select(PurchaseOrder).where(PurchaseOrder.id == poi.order_id)
        )
        order = order_result.scalar_one_or_none()
        if order:
            sup_name = order.supplier_name or "غير محدد"
            if sup_name not in supplier_prices or poi.unit_price < supplier_prices[sup_name]["price"]:
                supplier_prices[sup_name] = {
                    "price": poi.unit_price,
                    "source": "purchase_order",
                    "unit": poi.unit,
                    "currency": order.currency or "SAR",
                    "item_name": poi.name,
                    "order_number": order.order_number,
                    "order_date": order.created_at.strftime("%Y-%m-%d") if order.created_at else None
                }
    
    # Find better prices (excluding current supplier)
    better_prices = []
    for sup_name, data in supplier_prices.items():
        if supplier_name and sup_name.lower() == supplier_name.lower():
            continue  # Skip current supplier
        
        if unit_price > 0 and data["price"] < unit_price:
            savings = round(unit_price - data["price"], 2)
            savings_percent = round((savings / unit_price) * 100, 1)
            better_prices.append({
                "supplier_name": sup_name,
                "price": data["price"],
                "unit": data.get("unit", ""),
                "currency": data.get("currency", "SAR"),
                "source": data["source"],
                "savings": savings,
                "savings_percent": savings_percent,
                "item_name": data["item_name"]
            })
    
    # Sort by price
    better_prices.sort(key=lambda x: x["price"])
    
    has_better_price = len(better_prices) > 0
    
    return {
        "item_name": item_name,
        "current_price": unit_price,
        "current_supplier": supplier_name,
        "has_better_price": has_better_price,
        "better_options": better_prices[:5],  # Top 5 better prices
        "all_suppliers": [
            {
                "supplier_name": k,
                "price": v["price"],
                "unit": v.get("unit", ""),
                "source": v["source"]
            }
            for k, v in sorted(supplier_prices.items(), key=lambda x: x[1]["price"])
        ][:10]
    }


@pg_catalog_router.post("/price-catalog/quick-add")
async def quick_add_catalog_item(
    item: QuickCatalogAdd,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """
    إضافة صنف سريعة للكتالوج (من شاشة أمر الشراء)
    """
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بإضافة أصناف")
    
    # Check if item already exists with same name and supplier
    existing = await session.execute(
        select(PriceCatalogItem).where(
            PriceCatalogItem.name == item.name,
            PriceCatalogItem.supplier_name == item.supplier_name
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="هذا الصنف موجود مسبقاً لهذا المورد")
    
    new_item = PriceCatalogItem(
        id=str(uuid.uuid4()),
        name=item.name,
        unit=item.unit,
        price=item.price,
        currency=item.currency,
        supplier_name=item.supplier_name,
        is_active=True,
        created_by=current_user.id,
        created_by_name=current_user.name
    )
    
    session.add(new_item)
    await session.commit()
    
    return {
        "message": "تم إضافة الصنف للكتالوج بنجاح",
        "item": {
            "id": new_item.id,
            "name": new_item.name,
            "price": new_item.price,
            "unit": new_item.unit,
            "supplier_name": new_item.supplier_name
        }
    }


@pg_catalog_router.post("/item-aliases")
async def create_alias(
    alias_data: AliasCreate,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Create a new item alias"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بإضافة مسميات بديلة")
    
    # Check if alias already exists
    existing = await session.execute(
        select(ItemAlias).where(ItemAlias.alias_name == alias_data.alias_name)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="هذا المسمى موجود مسبقاً")
    
    # Get catalog item name
    catalog_result = await session.execute(
        select(PriceCatalogItem).where(PriceCatalogItem.id == alias_data.catalog_item_id)
    )
    catalog_item = catalog_result.scalar_one_or_none()
    
    if not catalog_item:
        raise HTTPException(status_code=404, detail="عنصر الكتالوج غير موجود")
    
    new_alias = ItemAlias(
        id=str(uuid.uuid4()),
        alias_name=alias_data.alias_name,
        catalog_item_id=alias_data.catalog_item_id,
        catalog_item_name=catalog_item.name,
        usage_count=0,
        created_by=current_user.id,
        created_by_name=current_user.name
    )
    
    session.add(new_alias)
    await session.commit()
    
    return {"message": "تم إضافة المسمى البديل بنجاح", "id": new_alias.id}


@pg_catalog_router.delete("/item-aliases/{alias_id}")
async def delete_alias(
    alias_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """Delete an item alias"""
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بحذف المسميات البديلة")
    
    result = await session.execute(
        select(ItemAlias).where(ItemAlias.id == alias_id)
    )
    alias = result.scalar_one_or_none()
    
    if not alias:
        raise HTTPException(status_code=404, detail="المسمى غير موجود")
    
    await session.delete(alias)
    await session.commit()
    
    return {"message": "تم حذف المسمى بنجاح"}
