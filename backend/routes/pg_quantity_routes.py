"""
PostgreSQL Quantity Engineer Routes - إدارة الكميات المخططة
النظام الصحيح: اختيار أصناف من كتالوج الأسعار وتحديد كميات مع تواريخ توريد
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timedelta
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func, desc, and_, or_
import uuid
import json
import io
import csv

from database import (
    get_postgres_session, User, Project, PlannedQuantity,
    PriceCatalogItem, BudgetCategory, PurchaseOrder, PurchaseOrderItem
)
from routes.pg_auth_routes import get_current_user_pg, UserRole

# Create router
pg_quantity_router = APIRouter(prefix="/api/pg/quantity", tags=["Quantity Engineer"])


# ==================== PYDANTIC MODELS ====================

class PlannedQuantityCreate(BaseModel):
    """إنشاء كمية مخططة - اختيار صنف من الكتالوج"""
    catalog_item_id: str  # معرف الصنف من كتالوج الأسعار (إجباري)
    project_id: str  # معرف المشروع (إجباري)
    planned_quantity: float  # الكمية المخططة
    expected_order_date: Optional[str] = None  # تاريخ الطلب المتوقع
    priority: int = 2  # الأولوية (1=عالية، 2=متوسطة، 3=منخفضة)
    notes: Optional[str] = None  # ملاحظات
    category_id: Optional[str] = None  # فئة الميزانية (اختياري - يمكن اختيارها يدوياً)


class PlannedQuantityUpdate(BaseModel):
    """تحديث كمية مخططة"""
    planned_quantity: Optional[float] = None
    expected_order_date: Optional[str] = None
    priority: Optional[int] = None
    notes: Optional[str] = None
    status: Optional[str] = None
    category_id: Optional[str] = None  # فئة الميزانية


class BulkPlannedQuantityCreate(BaseModel):
    """إضافة كميات مخططة بالجملة"""
    items: List[PlannedQuantityCreate]


class DeductQuantityRequest(BaseModel):
    """خصم كمية من الخطة عند إنشاء أمر شراء"""
    catalog_item_id: str
    project_id: str
    quantity_to_deduct: float
    order_id: Optional[str] = None


# ==================== HELPER FUNCTIONS ====================

def require_quantity_access(current_user: User):
    """التحقق من صلاحية الوصول لميزات مهندس الكميات"""
    allowed_roles = [
        UserRole.QUANTITY_ENGINEER,
        UserRole.PROCUREMENT_MANAGER,
        UserRole.SYSTEM_ADMIN
    ]
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")


# ==================== CATALOG ITEMS ENDPOINTS ====================

@pg_quantity_router.get("/catalog-items")
async def get_catalog_items_for_planning(
    search: Optional[str] = None,
    supplier_id: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """جلب أصناف كتالوج الأسعار للاختيار منها"""
    require_quantity_access(current_user)
    
    query = select(PriceCatalogItem).where(PriceCatalogItem.is_active == True)
    
    if search:
        query = query.where(
            or_(
                PriceCatalogItem.name.ilike(f"%{search}%"),
                PriceCatalogItem.description.ilike(f"%{search}%"),
                PriceCatalogItem.item_code.ilike(f"%{search}%")
            )
        )
    
    if supplier_id:
        query = query.where(PriceCatalogItem.supplier_id == supplier_id)
    
    # Count total
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await session.execute(count_query)
    total = total_result.scalar()
    
    # Pagination
    offset = (page - 1) * page_size
    query = query.order_by(PriceCatalogItem.item_code.asc().nullslast(), PriceCatalogItem.name).offset(offset).limit(page_size)
    
    result = await session.execute(query)
    items = result.scalars().all()
    
    return {
        "items": [
            {
                "id": item.id,
                "item_code": item.item_code,
                "name": item.name,
                "description": item.description,
                "unit": item.unit,
                "supplier_id": item.supplier_id,
                "supplier_name": item.supplier_name,
                "price": item.price,
                "currency": item.currency,
                "category_id": item.category_id,
                "category_name": item.category_name
            }
            for item in items
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@pg_quantity_router.get("/budget-categories/{project_id}")
async def get_budget_categories_for_project(
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """جلب فئات الميزانية لمشروع معين - لاختيارها عند إضافة كمية مخططة"""
    require_quantity_access(current_user)
    
    result = await session.execute(
        select(BudgetCategory).where(BudgetCategory.project_id == project_id).order_by(BudgetCategory.name)
    )
    categories = result.scalars().all()
    
    return [
        {
            "id": cat.id,
            "code": cat.code,
            "name": cat.name,
            "estimated_budget": cat.estimated_budget
        }
        for cat in categories
    ]


# ==================== PLANNED QUANTITIES CRUD ====================

@pg_quantity_router.get("/planned")
async def get_planned_quantities(
    project_id: Optional[str] = None,
    catalog_item_id: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """جلب الكميات المخططة مع الفلاتر"""
    require_quantity_access(current_user)
    
    query = select(PlannedQuantity)
    
    # Apply filters
    if project_id:
        query = query.where(PlannedQuantity.project_id == project_id)
    
    if catalog_item_id:
        query = query.where(PlannedQuantity.catalog_item_id == catalog_item_id)
    
    if status:
        query = query.where(PlannedQuantity.status == status)
    
    if search:
        query = query.where(
            or_(
                PlannedQuantity.item_name.ilike(f"%{search}%"),
                PlannedQuantity.item_code.ilike(f"%{search}%")
            )
        )
    
    # Count total
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await session.execute(count_query)
    total = total_result.scalar()
    
    # Pagination
    offset = (page - 1) * page_size
    query = query.order_by(desc(PlannedQuantity.created_at)).offset(offset).limit(page_size)
    
    result = await session.execute(query)
    items = result.scalars().all()
    
    return {
        "items": [
            {
                "id": item.id,
                "item_name": item.item_name,
                "item_code": item.item_code,
                "unit": item.unit,
                "description": item.description,
                "planned_quantity": item.planned_quantity,
                "ordered_quantity": item.ordered_quantity,
                "remaining_quantity": item.remaining_quantity,
                "project_id": item.project_id,
                "project_name": item.project_name,
                "category_id": item.category_id,
                "category_name": item.category_name,
                "catalog_item_id": item.catalog_item_id,
                "supplier_name": getattr(item, 'supplier_name', None),
                "unit_price": getattr(item, 'unit_price', None),
                "expected_order_date": item.expected_order_date.isoformat() if item.expected_order_date else None,
                "status": item.status,
                "priority": item.priority,
                "notes": item.notes,
                "created_by": item.created_by,
                "created_by_name": item.created_by_name,
                "created_at": item.created_at.isoformat() if item.created_at else None,
                "updated_at": item.updated_at.isoformat() if item.updated_at else None
            }
            for item in items
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@pg_quantity_router.post("/planned")
async def create_planned_quantity(
    data: PlannedQuantityCreate,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """إنشاء كمية مخططة جديدة - اختيار صنف من الكتالوج"""
    require_quantity_access(current_user)
    
    # جلب الصنف من الكتالوج
    catalog_result = await session.execute(
        select(PriceCatalogItem).where(PriceCatalogItem.id == data.catalog_item_id)
    )
    catalog_item = catalog_result.scalar_one_or_none()
    if not catalog_item:
        raise HTTPException(status_code=404, detail="الصنف غير موجود في الكتالوج")
    
    # جلب المشروع
    project_result = await session.execute(
        select(Project).where(Project.id == data.project_id)
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # جلب فئة الميزانية إذا تم تحديدها
    category_id = data.category_id or catalog_item.category_id
    category_name = None
    if category_id:
        cat_result = await session.execute(
            select(BudgetCategory).where(BudgetCategory.id == category_id)
        )
        category = cat_result.scalar_one_or_none()
        if category:
            category_name = category.name
        else:
            # استخدام category_name من الكتالوج إذا لم نجد الفئة
            category_name = catalog_item.category_name
    else:
        category_name = catalog_item.category_name
    
    # Parse expected order date
    expected_date = None
    if data.expected_order_date:
        try:
            expected_date = datetime.fromisoformat(data.expected_order_date.replace('Z', '+00:00'))
        except:
            try:
                expected_date = datetime.strptime(data.expected_order_date[:10], "%Y-%m-%d")
            except:
                pass
    
    new_item = PlannedQuantity(
        id=str(uuid.uuid4()),
        # بيانات الصنف من الكتالوج
        item_name=catalog_item.name,
        item_code=catalog_item.item_code or f"ITM{catalog_item.id[:5].upper()}",  # استخدام item_code من الكتالوج
        unit=catalog_item.unit,
        description=catalog_item.description,
        # الكميات
        planned_quantity=data.planned_quantity,
        ordered_quantity=0,
        remaining_quantity=data.planned_quantity,
        # المشروع
        project_id=data.project_id,
        project_name=project.name,
        # الفئة (من اختيار المستخدم أو من الكتالوج)
        category_id=category_id,
        category_name=category_name,
        # ربط بالكتالوج
        catalog_item_id=data.catalog_item_id,
        # تاريخ الطلب المتوقع
        expected_order_date=expected_date,
        # الحالة والأولوية
        status="planned",
        priority=data.priority,
        notes=data.notes,
        # التتبع
        created_by=current_user.id,
        created_by_name=current_user.name
    )
    
    session.add(new_item)
    await session.commit()
    
    return {
        "message": "تم إضافة الكمية المخططة بنجاح",
        "id": new_item.id,
        "item_name": new_item.item_name
    }


@pg_quantity_router.post("/planned/bulk")
async def create_bulk_planned_quantities(
    data: BulkPlannedQuantityCreate,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """إضافة كميات مخططة بالجملة"""
    require_quantity_access(current_user)
    
    created = 0
    errors = []
    
    for idx, item_data in enumerate(data.items):
        try:
            # جلب الصنف من الكتالوج
            catalog_result = await session.execute(
                select(PriceCatalogItem).where(PriceCatalogItem.id == item_data.catalog_item_id)
            )
            catalog_item = catalog_result.scalar_one_or_none()
            if not catalog_item:
                errors.append(f"العنصر {idx+1}: الصنف غير موجود في الكتالوج")
                continue
            
            # جلب المشروع
            project_result = await session.execute(
                select(Project).where(Project.id == item_data.project_id)
            )
            project = project_result.scalar_one_or_none()
            if not project:
                errors.append(f"العنصر {idx+1}: المشروع غير موجود")
                continue
            
            # Parse expected date
            expected_date = None
            if item_data.expected_order_date:
                try:
                    expected_date = datetime.fromisoformat(item_data.expected_order_date.replace('Z', '+00:00'))
                except:
                    try:
                        expected_date = datetime.strptime(item_data.expected_order_date[:10], "%Y-%m-%d")
                    except:
                        pass
            
            # جلب فئة الميزانية إذا تم تحديدها
            bulk_category_id = getattr(item_data, 'category_id', None) or catalog_item.category_id
            bulk_category_name = None
            if bulk_category_id:
                cat_result = await session.execute(
                    select(BudgetCategory).where(BudgetCategory.id == bulk_category_id)
                )
                cat = cat_result.scalar_one_or_none()
                bulk_category_name = cat.name if cat else catalog_item.category_name
            else:
                bulk_category_name = catalog_item.category_name
            
            new_item = PlannedQuantity(
                id=str(uuid.uuid4()),
                item_name=catalog_item.name,
                item_code=catalog_item.item_code or f"ITM{catalog_item.id[:5].upper()}",
                unit=catalog_item.unit,
                description=catalog_item.description,
                planned_quantity=item_data.planned_quantity,
                ordered_quantity=0,
                remaining_quantity=item_data.planned_quantity,
                project_id=item_data.project_id,
                project_name=project.name,
                category_id=bulk_category_id,
                category_name=bulk_category_name,
                catalog_item_id=item_data.catalog_item_id,
                expected_order_date=expected_date,
                status="planned",
                priority=item_data.priority,
                notes=item_data.notes,
                created_by=current_user.id,
                created_by_name=current_user.name
            )
            
            session.add(new_item)
            created += 1
            
        except Exception as e:
            errors.append(f"العنصر {idx+1}: {str(e)}")
    
    await session.commit()
    
    return {
        "message": f"تم إضافة {created} عنصر بنجاح",
        "created": created,
        "errors": errors[:10] if errors else []
    }


@pg_quantity_router.put("/planned/{item_id}")
async def update_planned_quantity(
    item_id: str,
    data: PlannedQuantityUpdate,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تحديث كمية مخططة"""
    require_quantity_access(current_user)
    
    result = await session.execute(
        select(PlannedQuantity).where(PlannedQuantity.id == item_id)
    )
    item = result.scalar_one_or_none()
    
    if not item:
        raise HTTPException(status_code=404, detail="العنصر غير موجود")
    
    # Update fields
    if data.planned_quantity is not None:
        old_planned = item.planned_quantity
        item.planned_quantity = data.planned_quantity
        # إعادة حساب الكمية المتبقية
        item.remaining_quantity = data.planned_quantity - item.ordered_quantity
        if item.remaining_quantity < 0:
            item.remaining_quantity = 0
        # تحديث الحالة
        if item.remaining_quantity == 0:
            item.status = "fully_ordered"
        elif item.ordered_quantity > 0:
            item.status = "partially_ordered"
        else:
            item.status = "planned"
    
    if data.priority is not None:
        item.priority = data.priority
    
    if data.notes is not None:
        item.notes = data.notes
    
    if data.status is not None:
        item.status = data.status
    
    if data.expected_order_date is not None:
        try:
            item.expected_order_date = datetime.fromisoformat(data.expected_order_date.replace('Z', '+00:00'))
        except:
            try:
                item.expected_order_date = datetime.strptime(data.expected_order_date[:10], "%Y-%m-%d")
            except:
                pass
    
    # تحديث فئة الميزانية
    if data.category_id is not None:
        if data.category_id == "":
            item.category_id = None
            item.category_name = None
        else:
            cat_result = await session.execute(
                select(BudgetCategory).where(BudgetCategory.id == data.category_id)
            )
            category = cat_result.scalar_one_or_none()
            if category:
                item.category_id = category.id
                item.category_name = category.name
    
    item.updated_at = datetime.utcnow()
    item.updated_by = current_user.id
    item.updated_by_name = current_user.name
    
    await session.commit()
    
    return {"message": "تم تحديث الكمية بنجاح"}


@pg_quantity_router.delete("/planned/{item_id}")
async def delete_planned_quantity(
    item_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """حذف كمية مخططة"""
    require_quantity_access(current_user)
    
    result = await session.execute(
        select(PlannedQuantity).where(PlannedQuantity.id == item_id)
    )
    item = result.scalar_one_or_none()
    
    if not item:
        raise HTTPException(status_code=404, detail="العنصر غير موجود")
    
    # منع الحذف إذا تم طلب جزء من الكمية
    if item.ordered_quantity > 0:
        raise HTTPException(status_code=400, detail="لا يمكن حذف عنصر تم طلبه بالفعل")
    
    await session.delete(item)
    await session.commit()
    
    return {"message": "تم حذف العنصر بنجاح"}


# ==================== خصم الكميات عند إنشاء أمر شراء ====================

@pg_quantity_router.post("/deduct")
async def deduct_quantity_from_plan(
    data: DeductQuantityRequest,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """خصم كمية من الخطة عند إنشاء أمر شراء"""
    # السماح لمدير المشتريات
    if current_user.role not in [UserRole.PROCUREMENT_MANAGER, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    # البحث عن الكميات المخططة لهذا الصنف في هذا المشروع
    result = await session.execute(
        select(PlannedQuantity).where(
            and_(
                PlannedQuantity.catalog_item_id == data.catalog_item_id,
                PlannedQuantity.project_id == data.project_id,
                PlannedQuantity.remaining_quantity > 0
            )
        ).order_by(PlannedQuantity.expected_order_date.asc().nullslast())
    )
    planned_items = result.scalars().all()
    
    if not planned_items:
        return {
            "message": "لا توجد كميات مخططة لهذا الصنف",
            "deducted": 0,
            "remaining_to_deduct": data.quantity_to_deduct
        }
    
    remaining_to_deduct = data.quantity_to_deduct
    total_deducted = 0
    
    for planned_item in planned_items:
        if remaining_to_deduct <= 0:
            break
        
        available = planned_item.remaining_quantity
        deduct_amount = min(available, remaining_to_deduct)
        
        planned_item.ordered_quantity += deduct_amount
        planned_item.remaining_quantity -= deduct_amount
        remaining_to_deduct -= deduct_amount
        total_deducted += deduct_amount
        
        # تحديث الحالة
        if planned_item.remaining_quantity == 0:
            planned_item.status = "fully_ordered"
        else:
            planned_item.status = "partially_ordered"
        
        planned_item.updated_at = datetime.utcnow()
    
    await session.commit()
    
    return {
        "message": f"تم خصم {total_deducted} من الكمية المخططة",
        "deducted": total_deducted,
        "remaining_to_deduct": remaining_to_deduct
    }


@pg_quantity_router.get("/check-planned/{catalog_item_id}/{project_id}")
async def check_planned_quantity(
    catalog_item_id: str,
    project_id: str,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """التحقق من الكمية المخططة لصنف معين في مشروع معين"""
    result = await session.execute(
        select(PlannedQuantity).where(
            and_(
                PlannedQuantity.catalog_item_id == catalog_item_id,
                PlannedQuantity.project_id == project_id
            )
        )
    )
    items = result.scalars().all()
    
    total_planned = sum(item.planned_quantity for item in items)
    total_ordered = sum(item.ordered_quantity for item in items)
    total_remaining = sum(item.remaining_quantity for item in items)
    
    return {
        "catalog_item_id": catalog_item_id,
        "project_id": project_id,
        "total_planned": total_planned,
        "total_ordered": total_ordered,
        "total_remaining": total_remaining,
        "entries": [
            {
                "id": item.id,
                "planned_quantity": item.planned_quantity,
                "ordered_quantity": item.ordered_quantity,
                "remaining_quantity": item.remaining_quantity,
                "expected_order_date": item.expected_order_date.isoformat() if item.expected_order_date else None,
                "status": item.status
            }
            for item in items
        ]
    }


# ==================== تقارير ولوحة المعلومات ====================

@pg_quantity_router.get("/dashboard/stats")
async def get_quantity_dashboard_stats(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """إحصائيات لوحة معلومات مهندس الكميات"""
    require_quantity_access(current_user)
    
    # جلب كل الكميات المخططة
    result = await session.execute(select(PlannedQuantity))
    items = result.scalars().all()
    
    # عدد المشاريع
    projects_result = await session.execute(select(func.count()).select_from(Project))
    projects_count = projects_result.scalar()
    
    # عدد أصناف الكتالوج النشطة
    catalog_result = await session.execute(
        select(func.count()).select_from(PriceCatalogItem).where(PriceCatalogItem.is_active == True)
    )
    catalog_count = catalog_result.scalar()
    
    now = datetime.utcnow()
    overdue = len([i for i in items if i.expected_order_date and i.expected_order_date < now and i.remaining_quantity > 0])
    due_soon = len([i for i in items if i.expected_order_date and now <= i.expected_order_date <= now + timedelta(days=10) and i.remaining_quantity > 0])
    
    return {
        "total_planned_items": len(items),
        "total_planned_qty": sum(i.planned_quantity for i in items),
        "total_ordered_qty": sum(i.ordered_quantity for i in items),
        "total_remaining_qty": sum(i.remaining_quantity for i in items),
        "overdue_items": overdue,
        "due_soon_items": due_soon,
        "projects_count": projects_count,
        "catalog_items_count": catalog_count
    }


@pg_quantity_router.get("/reports/summary")
async def get_quantity_summary_report(
    project_id: Optional[str] = None,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تقرير ملخص الكميات المخططة"""
    # السماح للمدراء بعرض التقارير
    allowed_roles = [
        UserRole.QUANTITY_ENGINEER,
        UserRole.PROCUREMENT_MANAGER,
        UserRole.GENERAL_MANAGER,
        UserRole.SYSTEM_ADMIN
    ]
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    query = select(PlannedQuantity)
    if project_id:
        query = query.where(PlannedQuantity.project_id == project_id)
    
    result = await session.execute(query)
    items = result.scalars().all()
    
    total_planned = sum(item.planned_quantity for item in items)
    total_ordered = sum(item.ordered_quantity for item in items)
    total_remaining = sum(item.remaining_quantity for item in items)
    
    # إحصاء حسب الحالة
    status_counts = {}
    for item in items:
        status_counts[item.status] = status_counts.get(item.status, 0) + 1
    
    now = datetime.utcnow()
    
    # الأصناف المتأخرة
    overdue_items = [
        item for item in items 
        if item.expected_order_date and item.expected_order_date < now and item.remaining_quantity > 0
    ]
    
    # الأصناف القريبة من الموعد (خلال 10 أيام)
    due_soon = [
        item for item in items
        if item.expected_order_date and now <= item.expected_order_date <= now + timedelta(days=10) and item.remaining_quantity > 0
    ]
    
    # حسب المشروع
    projects_data = {}
    for item in items:
        if item.project_name not in projects_data:
            projects_data[item.project_name] = {
                "project_name": item.project_name,
                "project_id": item.project_id,
                "total_items": 0,
                "planned_qty": 0,
                "ordered_qty": 0,
                "remaining_qty": 0
            }
        projects_data[item.project_name]["total_items"] += 1
        projects_data[item.project_name]["planned_qty"] += item.planned_quantity
        projects_data[item.project_name]["ordered_qty"] += item.ordered_quantity
        projects_data[item.project_name]["remaining_qty"] += item.remaining_quantity
    
    return {
        "summary": {
            "total_items": len(items),
            "total_planned_qty": total_planned,
            "total_ordered_qty": total_ordered,
            "total_remaining_qty": total_remaining,
            "completion_rate": round((total_ordered / total_planned * 100), 1) if total_planned > 0 else 0,
            "overdue_count": len(overdue_items),
            "due_soon_count": len(due_soon)
        },
        "status_breakdown": status_counts,
        "by_project": list(projects_data.values()),
        "overdue_items": [
            {
                "id": item.id,
                "item_name": item.item_name,
                "project_name": item.project_name,
                "remaining_qty": item.remaining_quantity,
                "expected_date": item.expected_order_date.isoformat() if item.expected_order_date else None,
                "days_overdue": (now - item.expected_order_date).days if item.expected_order_date else 0
            }
            for item in sorted(overdue_items, key=lambda x: x.expected_order_date or now)[:10]
        ],
        "due_soon_items": [
            {
                "id": item.id,
                "item_name": item.item_name,
                "project_name": item.project_name,
                "remaining_qty": item.remaining_quantity,
                "expected_date": item.expected_order_date.isoformat() if item.expected_order_date else None,
                "days_until": (item.expected_order_date - now).days if item.expected_order_date else 0
            }
            for item in sorted(due_soon, key=lambda x: x.expected_order_date or now)[:10]
        ]
    }


# ==================== تنبيهات المشرف ====================

@pg_quantity_router.get("/alerts")
async def get_supervisor_alerts(
    days_threshold: int = 7,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """جلب التنبيهات للمشرف - الأصناف المتأخرة والقريبة من الموعد"""
    allowed_roles = [
        UserRole.QUANTITY_ENGINEER,
        UserRole.PROCUREMENT_MANAGER,
        UserRole.GENERAL_MANAGER,
        UserRole.SUPERVISOR,
        UserRole.SYSTEM_ADMIN
    ]
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    now = datetime.utcnow()
    threshold_date = now + timedelta(days=days_threshold)
    
    # الأصناف المتأخرة
    overdue_result = await session.execute(
        select(PlannedQuantity).where(
            and_(
                PlannedQuantity.expected_order_date < now,
                PlannedQuantity.remaining_quantity > 0
            )
        ).order_by(PlannedQuantity.expected_order_date.asc())
    )
    overdue_items = overdue_result.scalars().all()
    
    # الأصناف القريبة من الموعد
    due_soon_result = await session.execute(
        select(PlannedQuantity).where(
            and_(
                PlannedQuantity.expected_order_date >= now,
                PlannedQuantity.expected_order_date <= threshold_date,
                PlannedQuantity.remaining_quantity > 0
            )
        ).order_by(PlannedQuantity.expected_order_date.asc())
    )
    due_soon_items = due_soon_result.scalars().all()
    
    # الأصناف ذات الأولوية العالية
    high_priority_result = await session.execute(
        select(PlannedQuantity).where(
            and_(
                PlannedQuantity.priority == 1,
                PlannedQuantity.remaining_quantity > 0
            )
        ).order_by(PlannedQuantity.expected_order_date.asc().nullslast())
    )
    high_priority_items = high_priority_result.scalars().all()
    
    return {
        "overdue": {
            "count": len(overdue_items),
            "items": [
                {
                    "id": item.id,
                    "item_name": item.item_name,
                    "project_name": item.project_name,
                    "remaining_qty": item.remaining_quantity,
                    "unit": item.unit,
                    "expected_date": item.expected_order_date.isoformat() if item.expected_order_date else None,
                    "days_overdue": (now - item.expected_order_date).days if item.expected_order_date else 0,
                    "priority": item.priority
                }
                for item in overdue_items[:20]
            ]
        },
        "due_soon": {
            "count": len(due_soon_items),
            "items": [
                {
                    "id": item.id,
                    "item_name": item.item_name,
                    "project_name": item.project_name,
                    "remaining_qty": item.remaining_quantity,
                    "unit": item.unit,
                    "expected_date": item.expected_order_date.isoformat() if item.expected_order_date else None,
                    "days_until": (item.expected_order_date - now).days if item.expected_order_date else 0,
                    "priority": item.priority
                }
                for item in due_soon_items[:20]
            ]
        },
        "high_priority": {
            "count": len(high_priority_items),
            "items": [
                {
                    "id": item.id,
                    "item_name": item.item_name,
                    "project_name": item.project_name,
                    "remaining_qty": item.remaining_quantity,
                    "unit": item.unit,
                    "expected_date": item.expected_order_date.isoformat() if item.expected_order_date else None,
                    "priority": item.priority
                }
                for item in high_priority_items[:20]
            ]
        }
    }


# ==================== تصدير البيانات ====================

@pg_quantity_router.get("/planned/export")
async def export_planned_quantities(
    project_id: Optional[str] = None,
    format: str = "excel",
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تصدير الكميات المخططة إلى Excel"""
    require_quantity_access(current_user)
    
    query = select(PlannedQuantity)
    if project_id:
        query = query.where(PlannedQuantity.project_id == project_id)
    query = query.order_by(desc(PlannedQuantity.created_at))
    
    result = await session.execute(query)
    items = result.scalars().all()
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة Excel غير متوفرة")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "الكميات المخططة"
    ws.sheet_view.rightToLeft = True
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['اسم الصنف', 'الوحدة', 'المشروع', 'الكمية المخططة', 'الكمية المطلوبة', 'الكمية المتبقية', 'تاريخ الطلب المتوقع', 'الحالة', 'الأولوية']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Data
    status_ar = {
        "planned": "مخطط",
        "partially_ordered": "طلب جزئي",
        "fully_ordered": "مكتمل",
        "overdue": "متأخر"
    }
    priority_ar = {1: "عالية", 2: "متوسطة", 3: "منخفضة"}
    
    for row_num, item in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=item.item_name).border = thin_border
        ws.cell(row=row_num, column=2, value=item.unit).border = thin_border
        ws.cell(row=row_num, column=3, value=item.project_name).border = thin_border
        ws.cell(row=row_num, column=4, value=item.planned_quantity).border = thin_border
        ws.cell(row=row_num, column=5, value=item.ordered_quantity).border = thin_border
        ws.cell(row=row_num, column=6, value=item.remaining_quantity).border = thin_border
        ws.cell(row=row_num, column=7, value=item.expected_order_date.strftime("%Y-%m-%d") if item.expected_order_date else "-").border = thin_border
        ws.cell(row=row_num, column=8, value=status_ar.get(item.status, item.status)).border = thin_border
        ws.cell(row=row_num, column=9, value=priority_ar.get(item.priority, "متوسطة")).border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=planned_quantities_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


# ==================== فئات الميزانية ====================

@pg_quantity_router.get("/budget-categories")
async def get_budget_categories(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """جلب فئات الميزانية للاختيار"""
    result = await session.execute(select(BudgetCategory))
    categories = result.scalars().all()
    
    return {
        "categories": [
            {
                "id": cat.id,
                "name": cat.name,
                "description": cat.description
            }
            for cat in categories
        ]
    }


# ==================== استيراد من Excel ====================

@pg_quantity_router.get("/planned/template")
async def download_planned_template(
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تحميل نموذج Excel للاستيراد - قسم الإدخال في الأعلى"""
    require_quantity_access(current_user)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة Excel غير متوفرة")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "الكميات المخططة"
    ws.sheet_view.rightToLeft = True
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    input_header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    item_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    project_header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    category_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    
    # جلب البيانات
    catalog_result = await session.execute(
        select(PriceCatalogItem).where(PriceCatalogItem.is_active == True).order_by(PriceCatalogItem.item_code.asc().nullslast())
    )
    catalog_items = catalog_result.scalars().all()
    
    projects_result = await session.execute(select(Project).order_by(Project.name))
    projects = projects_result.scalars().all()
    
    categories_result = await session.execute(select(BudgetCategory).order_by(BudgetCategory.code.asc().nullslast()))
    categories = categories_result.scalars().all()
    
    # === قسم الإدخال (في الأعلى) ===
    ws.merge_cells('A1:I1')
    input_title = ws['A1']
    input_title.value = "✏️ أدخل الكميات المخططة هنا (انسخ كود الصنف من القائمة أدناه)"
    input_title.font = Font(bold=True, size=14, color="FFFFFF")
    input_title.fill = input_header_fill
    input_title.alignment = center_align
    
    # رؤوس قسم الإدخال - الترتيب الجديد
    input_headers = ['كود الصنف *', 'اسم الصنف *', 'كود التصنيف', 'تصنيف الميزانية', 'اسم المشروع *', 'الكمية المخططة *', 'تاريخ الطلب المتوقع', 'الأولوية (1-3)', 'ملاحظات']
    for col, header in enumerate(input_headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    # صف مثال
    if catalog_items and projects:
        example_item = catalog_items[0]
        example_project = projects[0]
        example_category = categories[0] if categories else None
        
        example_data = [
            example_item.item_code or f"ITM{example_item.id[:5].upper()}",
            example_item.name,
            example_category.code if example_category else "",
            example_category.name if example_category else "",
            example_project.name,
            25,
            "2026-02-01",
            2,
            "مثال - احذف هذا الصف"
        ]
        
        for col, value in enumerate(example_data, 1):
            cell = ws.cell(row=3, column=col, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = right_align
    
    # صفوف فارغة للإدخال
    for row in range(4, 14):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = thin_border
    
    # === قسم الأصناف المتاحة ===
    items_start_row = 16
    
    ws.merge_cells(f'A{items_start_row}:F{items_start_row}')
    items_title = ws[f'A{items_start_row}']
    items_title.value = "📋 الأصناف المتاحة للاستيراد"
    items_title.font = Font(bold=True, size=14, color="FFFFFF")
    items_title.fill = item_header_fill
    items_title.alignment = center_align
    
    item_headers = ['كود الصنف', 'اسم الصنف', 'الوحدة', 'السعر', 'المورد', 'التصنيف']
    for col, header in enumerate(item_headers, 1):
        cell = ws.cell(row=items_start_row + 1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    for row_num, item in enumerate(catalog_items, items_start_row + 2):
        ws.cell(row=row_num, column=1, value=item.item_code or f"ITM{item.id[:5].upper()}").border = thin_border
        ws.cell(row=row_num, column=2, value=item.name).border = thin_border
        ws.cell(row=row_num, column=3, value=item.unit).border = thin_border
        ws.cell(row=row_num, column=4, value=item.price).border = thin_border
        ws.cell(row=row_num, column=5, value=item.supplier_name or "-").border = thin_border
        ws.cell(row=row_num, column=6, value=item.category_name or "-").border = thin_border
    
    # === قسم المشاريع ===
    projects_start_row = items_start_row + len(catalog_items) + 4
    
    ws.merge_cells(f'A{projects_start_row}:B{projects_start_row}')
    projects_title = ws[f'A{projects_start_row}']
    projects_title.value = "📁 المشاريع المتاحة"
    projects_title.font = Font(bold=True, size=14, color="FFFFFF")
    projects_title.fill = project_header_fill
    projects_title.alignment = center_align
    
    project_headers = ['اسم المشروع', 'كود المشروع']
    for col, header in enumerate(project_headers, 1):
        cell = ws.cell(row=projects_start_row + 1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    for row_num, project in enumerate(projects, projects_start_row + 2):
        ws.cell(row=row_num, column=1, value=project.name).border = thin_border
        ws.cell(row=row_num, column=2, value=getattr(project, 'code', None) or "-").border = thin_border
    
    # === قسم فئات الميزانية ===
    categories_start_row = projects_start_row + len(projects) + 4
    
    ws.merge_cells(f'A{categories_start_row}:C{categories_start_row}')
    categories_title = ws[f'A{categories_start_row}']
    categories_title.value = "🏷️ فئات الميزانية المتاحة"
    categories_title.font = Font(bold=True, size=14, color="FFFFFF")
    categories_title.fill = category_header_fill
    categories_title.alignment = center_align
    
    cat_headers = ['كود الفئة', 'اسم الفئة', 'المشروع']
    for col, header in enumerate(cat_headers, 1):
        cell = ws.cell(row=categories_start_row + 1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    for row_num, cat in enumerate(categories, categories_start_row + 2):
        ws.cell(row=row_num, column=1, value=cat.code or "-").border = thin_border
        ws.cell(row=row_num, column=2, value=cat.name).border = thin_border
        ws.cell(row=row_num, column=3, value=cat.project_name or "-").border = thin_border
    
    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=planned_quantities_template.xlsx"}
    )


@pg_quantity_router.post("/planned/import")
async def import_planned_quantities(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """استيراد الكميات المخططة من Excel - الترتيب الجديد"""
    require_quantity_access(current_user)
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="يجب أن يكون الملف بصيغة Excel (.xlsx أو .xls)")
    
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة Excel غير متوفرة")
    
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    created = 0
    errors = []
    
    # البحث عن صف بداية الإدخال (يحتوي على "كود الصنف")
    start_row = 2  # افتراضياً - بعد رؤوس الأعمدة
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if row and row[0] and 'كود الصنف' in str(row[0]):
            start_row = row_num + 1
            break
    
    # الترتيب الجديد للأعمدة:
    # 0: كود الصنف | 1: اسم الصنف | 2: كود التصنيف | 3: تصنيف الميزانية | 4: اسم المشروع | 5: الكمية | 6: التاريخ | 7: الأولوية | 8: ملاحظات
    
    for row_num, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start_row):
        # تجاهل الصفوف الفارغة
        if not row or not row[0]:
            continue
        
        # تجاهل صفوف العناوين والأقسام
        first_cell = str(row[0]).strip()
        if first_cell.startswith('📋') or first_cell.startswith('📁') or first_cell.startswith('✏️') or first_cell.startswith('🏷️'):
            break  # وصلنا لقسم البيانات المرجعية - توقف
        if 'الصنف' in first_cell or 'المشروع' in first_cell or 'اسم' in first_cell or 'الفئة' in first_cell:
            continue
        
        # استخراج البيانات من الأعمدة
        item_code = first_cell
        item_name = str(row[1]).strip() if len(row) > 1 and row[1] else None
        category_code = str(row[2]).strip() if len(row) > 2 and row[2] else None
        category_name_input = str(row[3]).strip() if len(row) > 3 and row[3] else None
        project_name = str(row[4]).strip() if len(row) > 4 and row[4] else None
        
        # الكمية في العمود السادس (index 5)
        try:
            planned_quantity = float(row[5]) if len(row) > 5 and row[5] else 0
        except (ValueError, TypeError):
            errors.append(f"صف {row_num}: الكمية '{row[5] if len(row) > 5 else 'فارغ'}' غير صالحة")
            continue
        
        if planned_quantity <= 0:
            continue
        
        # التحقق من البيانات المطلوبة
        if not project_name:
            errors.append(f"صف {row_num}: اسم المشروع مطلوب")
            continue
        
        expected_order_date = None
        priority = 2
        notes = None
        
        # معالجة التاريخ (العمود السابع - index 6)
        if len(row) > 6 and row[6]:
            try:
                if hasattr(row[6], 'strftime'):
                    expected_order_date = row[6]
                else:
                    expected_order_date = datetime.strptime(str(row[6])[:10], "%Y-%m-%d")
            except:
                pass
        
        # معالجة الأولوية (العمود الثامن - index 7)
        if len(row) > 7 and row[7]:
            try:
                priority = int(row[7])
                if priority not in [1, 2, 3]:
                    priority = 2
            except (ValueError, TypeError):
                priority = 2
        
        # معالجة الملاحظات (العمود التاسع - index 8)
        if len(row) > 8 and row[8]:
            notes = str(row[8]).strip()
            if notes.lower() in ['none', 'مثال', 'مثال - احذف هذا الصف']:
                continue  # تجاهل صف المثال
        
        # البحث عن الصنف بالكود أولاً
        catalog_result = await session.execute(
            select(PriceCatalogItem).where(
                or_(
                    PriceCatalogItem.item_code == item_code,
                    PriceCatalogItem.id == item_code,
                    PriceCatalogItem.name == item_code
                )
            )
        )
        catalog_item = catalog_result.scalar_one_or_none()
        if not catalog_item:
            errors.append(f"صف {row_num}: الصنف '{item_code}' غير موجود في الكتالوج")
            continue
        
        # البحث عن المشروع بالاسم أو الكود
        project_result = await session.execute(
            select(Project).where(
                or_(
                    Project.name == project_name,
                    Project.id == project_name,
                    func.lower(Project.name) == func.lower(project_name),
                    Project.code == project_name
                )
            )
        )
        project = project_result.scalar_one_or_none()
        if not project:
            errors.append(f"صف {row_num}: المشروع '{project_name}' غير موجود")
            continue
        
        # البحث عن فئة الميزانية بالكود أو الاسم
        import_category_id = None
        import_category_name = None
        
        # أولاً: البحث بكود الفئة
        if category_code and category_code != '-' and category_code.lower() != 'none':
            cat_result = await session.execute(
                select(BudgetCategory).where(
                    or_(
                        BudgetCategory.code == category_code,
                        BudgetCategory.name == category_code,
                        BudgetCategory.id == category_code
                    )
                )
            )
            category = cat_result.scalar_one_or_none()
            if category:
                import_category_id = category.id
                import_category_name = category.name
        
        # ثانياً: البحث باسم الفئة إذا لم نجد بالكود
        if not import_category_id and category_name_input and category_name_input != '-' and category_name_input.lower() != 'none':
            cat_result = await session.execute(
                select(BudgetCategory).where(
                    or_(
                        BudgetCategory.name == category_name_input,
                        func.lower(BudgetCategory.name) == func.lower(category_name_input)
                    )
                )
            )
            category = cat_result.scalar_one_or_none()
            if category:
                import_category_id = category.id
                import_category_name = category.name
        
        # استخدام الفئة من الاستيراد أو من الكتالوج
        final_category_id = import_category_id or catalog_item.category_id
        final_category_name = import_category_name or catalog_item.category_name
        
        new_item = PlannedQuantity(
            id=str(uuid.uuid4()),
            item_name=catalog_item.name,
            item_code=catalog_item.item_code or f"ITM{catalog_item.id[:5].upper()}",
            unit=catalog_item.unit,
            description=catalog_item.description,
            planned_quantity=planned_quantity,
            ordered_quantity=0,
            remaining_quantity=planned_quantity,
            project_id=project.id,
            project_name=project.name,
            category_id=final_category_id,
            category_name=final_category_name,
            catalog_item_id=catalog_item.id,
            expected_order_date=expected_order_date,
            status="planned",
            priority=priority,
            notes=notes,
            created_by=current_user.id,
            created_by_name=current_user.name
        )
        
        session.add(new_item)
        created += 1
    
    await session.commit()
    
    return {
        "message": f"تم استيراد {created} كمية مخططة بنجاح",
        "created": created,
        "errors": errors[:20]
    }


# ==================== تصدير التقارير ====================

@pg_quantity_router.get("/reports/export")
async def export_quantity_report(
    project_id: Optional[str] = None,
    format: str = "excel",
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """تصدير تقرير الكميات إلى Excel أو PDF"""
    allowed_roles = [
        UserRole.QUANTITY_ENGINEER,
        UserRole.PROCUREMENT_MANAGER,
        UserRole.GENERAL_MANAGER,
        UserRole.SYSTEM_ADMIN
    ]
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    query = select(PlannedQuantity)
    if project_id:
        query = query.where(PlannedQuantity.project_id == project_id)
    query = query.order_by(PlannedQuantity.project_name, desc(PlannedQuantity.created_at))
    
    result = await session.execute(query)
    items = result.scalars().all()
    
    now = datetime.utcnow()
    
    # Calculate summary
    total_planned = sum(item.planned_quantity for item in items)
    total_ordered = sum(item.ordered_quantity for item in items)
    total_remaining = sum(item.remaining_quantity for item in items)
    overdue_count = len([i for i in items if i.expected_order_date and i.expected_order_date < now and i.remaining_quantity > 0])
    
    # Group by project
    projects_data = {}
    for item in items:
        if item.project_name not in projects_data:
            projects_data[item.project_name] = {
                "items": [],
                "planned_qty": 0,
                "ordered_qty": 0,
                "remaining_qty": 0
            }
        projects_data[item.project_name]["items"].append(item)
        projects_data[item.project_name]["planned_qty"] += item.planned_quantity
        projects_data[item.project_name]["ordered_qty"] += item.ordered_quantity
        projects_data[item.project_name]["remaining_qty"] += item.remaining_quantity
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة Excel غير متوفرة")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير الكميات"
    ws.sheet_view.rightToLeft = True
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    summary_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    project_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws.cell(row=1, column=1, value="تقرير الكميات المخططة").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    # Summary
    row = 4
    ws.cell(row=row, column=1, value="ملخص التقرير").font = Font(bold=True)
    row += 1
    summary_data = [
        ("إجمالي الأصناف", len(items)),
        ("الكمية المخططة", total_planned),
        ("الكمية المطلوبة", total_ordered),
        ("الكمية المتبقية", total_remaining),
        ("نسبة الإنجاز", f"{round((total_ordered / total_planned * 100), 1) if total_planned > 0 else 0}%"),
        ("الأصناف المتأخرة", overdue_count)
    ]
    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).fill = summary_fill
        ws.cell(row=row, column=2, value=value).fill = summary_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).border = thin_border
        row += 1
    
    # Details by project
    row += 2
    status_ar = {"planned": "مخطط", "partially_ordered": "طلب جزئي", "fully_ordered": "مكتمل"}
    priority_ar = {1: "عالية", 2: "متوسطة", 3: "منخفضة"}
    
    for project_name, project_data in projects_data.items():
        # Project header
        ws.cell(row=row, column=1, value=f"المشروع: {project_name}").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = project_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1
        
        ws.cell(row=row, column=1, value=f"إجمالي: مخطط {project_data['planned_qty']} | مطلوب {project_data['ordered_qty']} | متبقي {project_data['remaining_qty']}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1
        
        # Headers
        headers = ['الصنف', 'الوحدة', 'التصنيف', 'مخطط', 'مطلوب', 'متبقي', 'تاريخ الطلب', 'الحالة']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        
        # Items
        for item in project_data["items"]:
            ws.cell(row=row, column=1, value=item.item_name).border = thin_border
            ws.cell(row=row, column=2, value=item.unit).border = thin_border
            ws.cell(row=row, column=3, value=item.category_name or "-").border = thin_border
            ws.cell(row=row, column=4, value=item.planned_quantity).border = thin_border
            ws.cell(row=row, column=5, value=item.ordered_quantity).border = thin_border
            ws.cell(row=row, column=6, value=item.remaining_quantity).border = thin_border
            ws.cell(row=row, column=7, value=item.expected_order_date.strftime("%Y-%m-%d") if item.expected_order_date else "-").border = thin_border
            ws.cell(row=row, column=8, value=status_ar.get(item.status, item.status)).border = thin_border
            row += 1
        
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=quantity_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


# ==================== APIs للمشرفين والمهندسين ====================

@pg_quantity_router.get("/by-role")
async def get_quantities_by_role(
    project_id: Optional[str] = None,
    current_user: User = Depends(get_current_user_pg),
    session: AsyncSession = Depends(get_postgres_session)
):
    """جلب الكميات المخططة حسب دور المستخدم"""
    allowed_roles = [
        UserRole.SUPERVISOR,
        UserRole.ENGINEER,
        UserRole.GENERAL_MANAGER,
        UserRole.PROCUREMENT_MANAGER,
        UserRole.SYSTEM_ADMIN
    ]
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="غير مصرح لك بهذا الإجراء")
    
    query = select(PlannedQuantity)
    
    # فلترة حسب المشروع
    if project_id:
        query = query.where(PlannedQuantity.project_id == project_id)
    
    query = query.order_by(desc(PlannedQuantity.created_at))
    
    result = await session.execute(query)
    items = result.scalars().all()
    
    now = datetime.utcnow()
    
    return {
        "items": [
            {
                "id": item.id,
                "item_name": item.item_name,
                "unit": item.unit,
                "category_name": item.category_name,
                "planned_quantity": item.planned_quantity,
                "ordered_quantity": item.ordered_quantity,
                "remaining_quantity": item.remaining_quantity,
                "project_id": item.project_id,
                "project_name": item.project_name,
                "expected_order_date": item.expected_order_date.isoformat() if item.expected_order_date else None,
                "status": item.status,
                "priority": item.priority,
                "is_overdue": item.expected_order_date and item.expected_order_date < now and item.remaining_quantity > 0,
                "days_until": (item.expected_order_date - now).days if item.expected_order_date and item.expected_order_date >= now else None
            }
            for item in items
        ],
        "summary": {
            "total_items": len(items),
            "total_planned": sum(i.planned_quantity for i in items),
            "total_ordered": sum(i.ordered_quantity for i in items),
            "total_remaining": sum(i.remaining_quantity for i in items),
            "overdue_count": len([i for i in items if i.expected_order_date and i.expected_order_date < now and i.remaining_quantity > 0])
        }
    }
